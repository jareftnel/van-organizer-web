#!/usr/bin/env python3
"""
Build Van Organizer HTML (v14 behavior) from:
- Excel: bag→overflow mapping (Bags_with_Overflow)
- PDF : Sort Zone + Bag Pkgs + wave time (merged by bag index)

Optimizations (no output/UX changes):
- Workbook opened in read_only mode
- PDF opened once (title + parsing in one pass)
- Optional on-disk cache for PDF parse (huge speedup on repeat runs)
"""

from __future__ import annotations
import argparse
import datetime as _dt
import json
import os
import re
from pathlib import Path
from typing import Dict, List, Tuple, Optional

import openpyxl
import pdfplumber


CACHE_VERSION_PDF = 3
CACHE_VERSION_ROUTES = 3

# ----------------------------- Regex (precompiled) -----------------------------
PAT_HEADER = re.compile(r'\b(DDF\d+)\s*·\s*([A-Z]{3},\s*[A-Z]{3}\s+\d{1,2},\s+\d{4})\b')
PAT_DATE_ONLY = re.compile(r'\b([A-Z]{3},\s*[A-Z]{3}\s+\d{1,2},\s+\d{4})\b')
PAT_FILE_DATE = re.compile(r'(\d{2})_(\d{2})_(\d{4})')

PAT_ROW_FULL = re.compile(r'^\s*(\d+)\s+([A-Z]-\d+(?:\.\d+)?[A-Z]?)\s+([A-Za-z]+)\s+([0-9A-Za-z]+)\s+(\d+)(?:\s+|$)')
PAT_ROW_NOSZ = re.compile(r'^\s*(\d+)\s+([A-Za-z]+)\s+([0-9A-Za-z]+)\s+(\d+)(?:\s+|$)')
PAT_TIME = re.compile(r'\b(\d{1,2}:\d{2}\s*[AP]M)\b')

PAT_OV_ZONE_CNT = re.compile(r'^([0-9]+\.[0-9]+[A-Z])\s*\((\d+)\)\s*$')
PAT_OV_ZONE = re.compile(r'^([0-9]+\.[0-9]+[A-Z])')
SHEET_RE = re.compile(r'^([A-Z]\.\d+)_?(CX\d+)$')


# ----------------------------- Helpers -----------------------------
def _time_to_minutes(t: str) -> Optional[int]:
    t = (t or "").strip().upper()
    m = re.match(r'^(\d{1,2}):(\d{2})\s*([AP]M)$', t)
    if not m:
        return None
    hh = int(m.group(1))
    mm = int(m.group(2))
    ap = m.group(3)
    if ap == "AM":
        if hh == 12:
            hh = 0
    else:
        if hh != 12:
            hh += 12
    return hh * 60 + mm


def _sort_route_short(rs: str) -> Tuple[str, int]:
    m = re.match(r'^([A-Z]+)\.(\d+)$', rs or "")
    if not m:
        return (rs or "", 0)
    return (m.group(1), int(m.group(2)))


def _parse_zone_counts(zones_str: str) -> List[Tuple[str, int]]:
    if not zones_str:
        return []
    parts = [p.strip() for p in str(zones_str).split(";") if p and str(p).strip()]
    out: List[Tuple[str, int]] = []
    for p in parts:
        m = PAT_OV_ZONE_CNT.match(p)
        if m:
            out.append((m.group(1), int(m.group(2))))
            continue
        m2 = PAT_OV_ZONE.match(p)
        if m2:
            out.append((m2.group(1), 0))
    return out


def _cache_path_for(pdf_path: str) -> Path:
    p = Path(pdf_path)
    return p.with_suffix(p.suffix + ".vanorg_cache.json")


def _load_pdf_cache(pdf_path: str) -> Optional[dict]:
    cache_path = _cache_path_for(pdf_path)
    if not cache_path.exists():
        return None
    try:
        st = os.stat(pdf_path)
        with cache_path.open("r", encoding="utf-8") as f:
            obj = json.load(f)
        meta = obj.get("meta", {})
        if meta.get("v") != CACHE_VERSION_PDF:
            return None
        if meta.get("size") != st.st_size:
            return None
        if meta.get("mtime") != int(st.st_mtime):
            return None
        data = obj.get("data")
        if not data or not data.get("pdf_meta"):
            return None
        return data
    except Exception:
        return None


def _save_pdf_cache(pdf_path: str, data: dict) -> None:
    try:
        st = os.stat(pdf_path)
        cache_path = _cache_path_for(pdf_path)
        payload = {
            "meta": {"v": CACHE_VERSION_PDF, "size": st.st_size, "mtime": int(st.st_mtime)},
            "data": data,
        }
        cache_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    except Exception:
        pass
def _routes_cache_path_for(xlsx_path: str) -> Path:
    p = Path(xlsx_path)
    return p.with_suffix(p.suffix + ".vanorg_routes_cache.json")


def _load_routes_cache(pdf_path: str, xlsx_path: str) -> Optional[dict]:
    cache_path = _routes_cache_path_for(xlsx_path)
    if not cache_path.exists():
        return None
    try:
        pst = os.stat(pdf_path)
        xst = os.stat(xlsx_path)
        with cache_path.open("r", encoding="utf-8") as f:
            obj = json.load(f)
        meta = obj.get("meta", {})
        if meta.get("v") != CACHE_VERSION_ROUTES:
            return None
        if meta.get("pdf_size") != pst.st_size or meta.get("pdf_mtime") != int(pst.st_mtime):
            return None
        if meta.get("xlsx_size") != xst.st_size or meta.get("xlsx_mtime") != int(xst.st_mtime):
            return None
        data = obj.get("data")
        if not data or not data.get("routes"):
            return None
        return data
    except Exception:
        return None


def _save_routes_cache(pdf_path: str, xlsx_path: str, data: dict) -> None:
    try:
        pst = os.stat(pdf_path)
        xst = os.stat(xlsx_path)
        cache_path = _routes_cache_path_for(xlsx_path)
        payload = {
            "meta": {
                "v": CACHE_VERSION_ROUTES,
                "pdf_size": pst.st_size, "pdf_mtime": int(pst.st_mtime),
                "xlsx_size": xst.st_size, "xlsx_mtime": int(xst.st_mtime),
            },
            "data": data,
        }
        cache_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    except Exception:
        pass



# ----------------------------- Parsing -----------------------------
def _extract_pkg_summaries(lines: List[str]) -> Tuple[Optional[int], Optional[int]]:
    commercial = None
    total = None
    for line in lines:
        s = line.strip().lower()
        if s.startswith("commercial packages"):
            for tok in reversed(line.split()):
                digits = re.sub(r"[^\d]", "", tok)
                if digits:
                    commercial = int(digits)
                    break
        if s.startswith("total packages"):
            for tok in reversed(line.split()):
                digits = re.sub(r"[^\d]", "", tok)
                if digits:
                    total = int(digits)
                    break
    return commercial, total


def parse_pdf_meta(
    pdf_path: str,
    use_cache: bool = True,
) -> Tuple[str, str, Dict[str, Dict[int, dict]], Dict[str, str], Dict[str, dict]]:
    """
    Returns:
      header_title, route_code, pdf_meta[route_short][idx] = {sort_zone, pkgs},
      route_time[route_short] = "11:20 AM", pkg_summary[route_short] = {commercial, total}
    """
    if use_cache:
        cached = _load_pdf_cache(pdf_path)
        if cached:
            return (
                cached["header_title"],
                cached["route_code"],
                cached["pdf_meta"],
                cached["route_time"],
                cached.get("pkg_summary") or {},
            )

    header_title = ""
    route_code = "DDF5"
    date_str = ""

    pdf_meta: Dict[str, Dict[int, dict]] = {}
    route_time: Dict[str, str] = {}
    pkg_summary: Dict[str, dict] = {}

    def _group_words_into_lines(words: List[dict], y_tol: float = 2.0) -> List[str]:
        if not words:
            return []
        words = sorted(words, key=lambda w: (float(w.get("top", 0.0)), float(w.get("x0", 0.0))))
        lines: List[List[dict]] = []
        cur: List[dict] = []
        cur_y: Optional[float] = None

        for w in words:
            y = float(w.get("top", 0.0))
            if cur_y is None:
                cur_y = y
                cur = [w]
                continue
            if abs(y - cur_y) <= y_tol:
                cur.append(w)
            else:
                lines.append(cur)
                cur = [w]
                cur_y = y
        if cur:
            lines.append(cur)

        out_lines: List[str] = []
        for ln in lines:
            ln_sorted = sorted(ln, key=lambda w: float(w.get("x0", 0.0)))
            text = " ".join((w.get("text") or "").strip() for w in ln_sorted if (w.get("text") or "").strip())
            text = re.sub(r"\s+", " ", text).strip()
            if text:
                out_lines.append(text)
        return out_lines

    with pdfplumber.open(pdf_path) as pdf:
        t0 = (pdf.pages[0].extract_text() or "")
        m = PAT_HEADER.search(t0)
        if m:
            route_code = m.group(1)
            date_str = m.group(2).upper()
        else:
            m2 = PAT_DATE_ONLY.search(t0)
            if m2:
                date_str = m2.group(1).upper()
            else:
                m3 = PAT_FILE_DATE.search(pdf_path)
                if m3:
                    mm, dd, yyyy = map(int, m3.groups())
                    dt = _dt.date(yyyy, mm, dd)
                    date_str = dt.strftime("%a, %b %d, %Y").upper()
        header_title = f"{route_code} • {date_str}".strip(" •")

        for page in pdf.pages:
            page_text = page.extract_text() or ""
            if "Sort Zone" not in page_text or "Pkgs" not in page_text:
                continue

            lines_quick = [ln.strip() for ln in page_text.splitlines() if ln and ln.strip()]
            route_short = ""
            for ln in lines_quick[:20]:
                if ln.startswith("STG."):
                    route_short = ln.replace("STG.", "").strip()
                    break
            if not route_short:
                words0 = page.extract_words(use_text_flow=True, keep_blank_chars=False) or []
                for w in words0:
                    t = (w.get("text") or "").strip()
                    if t.startswith("STG."):
                        route_short = t.replace("STG.", "").strip()
                        break
            if not route_short:
                continue

            comm_pkgs, total_pkgs = _extract_pkg_summaries(lines_quick)
            if comm_pkgs is not None or total_pkgs is not None:
                summary = pkg_summary.get(route_short) or {}
                if comm_pkgs is not None:
                    summary["commercial"] = comm_pkgs
                if total_pkgs is not None:
                    summary["total"] = total_pkgs
                pkg_summary[route_short] = summary

            if route_short not in route_time:
                tm = PAT_TIME.search(page_text)
                if tm:
                    route_time[route_short] = tm.group(1).upper()

            meta_by_idx = pdf_meta.get(route_short)
            if meta_by_idx is None:
                meta_by_idx = {}
                pdf_meta[route_short] = meta_by_idx

            words = page.extract_words(use_text_flow=True, keep_blank_chars=False) or []
            line_texts = _group_words_into_lines(words, y_tol=2.0)

            for ln in line_texts:
                m = PAT_ROW_FULL.match(ln)
                if m:
                    idx = int(m.group(1))
                    meta_by_idx[idx] = {"sort_zone": m.group(2), "pkgs": int(m.group(5))}
                    continue
                m2 = PAT_ROW_NOSZ.match(ln)
                if m2:
                    idx = int(m2.group(1))
                    if idx not in meta_by_idx:
                        meta_by_idx[idx] = {"sort_zone": "", "pkgs": int(m2.group(4))}

    if use_cache and pdf_meta:
        _save_pdf_cache(pdf_path, {
            "header_title": header_title,
            "route_code": route_code,
            "pdf_meta": pdf_meta,
            "route_time": route_time,
            "pkg_summary": pkg_summary,
        })

    return header_title, route_code, pdf_meta, route_time, pkg_summary


def parse_excel_routes(
    xlsx_path: str,
    pdf_meta: Dict[str, Dict[int, dict]],
    route_time: Dict[str, str],
    pkg_summary: Dict[str, dict],
) -> List[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)

    routes: List[dict] = []
    for sheet_name in wb.sheetnames:
        if sheet_name == "INDEX":
            continue
        m = SHEET_RE.match(sheet_name)
        if not m:
            continue
        rs, cx = m.group(1), m.group(2)
        ws = wb[sheet_name]

        combined: List[dict] = []
        bags: List[str] = []
        overflow_total = 0
        ov_agg: Dict[str, int] = {}

        ov_seq: List[dict] = []
        # Rows: Bag | Overflow Zone(s) | Overflow Pkgs (total)
        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            bag = row[0] if len(row) > 0 else None
            zones = row[1] if len(row) > 1 else None
            total_cell = row[2] if len(row) > 2 else None
            if bag is None:
                continue
            bag_s = str(bag).strip()
            if not bag_s:
                continue

            zones_s = "" if zones is None else str(zones).strip()

            total_val = None
            if total_cell is not None:
                tc_s = str(total_cell).strip()
                if tc_s:
                    # Excel might store as float
                    total_val = int(float(total_cell))

            zone_counts = _parse_zone_counts(zones_s)
            zone_total = sum(cnt for _, cnt in zone_counts)
            if total_val is None and zone_total:
                total_val = zone_total

            combined.append({"bag": bag_s, "zones": zones_s, "total": "" if total_val is None else str(total_val)})
            bags.append(bag_s)

            if total_val is not None:
                overflow_total += total_val

            for z, cnt in zone_counts:
                ov_seq.append({"zone": z, "count": cnt, "bag_idx": len(bags)})
                ov_agg[z] = ov_agg.get(z, 0) + cnt

        overflow_agg = [{"zone": k, "count": v} for k, v in sorted(ov_agg.items(), key=lambda x: x[0])]

        meta_by_idx = pdf_meta.get(rs, {})
        bags_detail: List[dict] = []
        for i, bag in enumerate(bags, start=1):
            bag_id = bag.split(" ", 1)[1] if " " in bag else bag
            meta = meta_by_idx.get(i)
            bags_detail.append({
                "idx": i,
                "bag": bag,
                "bag_id": bag_id,
                "sort_zone": meta["sort_zone"] if meta else "",
                "pkgs": meta["pkgs"] if meta else None
            })

        pkg_info = pkg_summary.get(rs) or {}
        total_pkgs = pkg_info.get("total")
        if total_pkgs is None:
            total_calc = sum(x["pkgs"] for x in bags_detail if x.get("pkgs") is not None)
            total_calc += overflow_total
            total_pkgs = total_calc if total_calc > 0 else None

        routes.append({
            "route_short": rs,
            "cx": cx,
            "wave_time": route_time.get(rs, ""),
            "bags_count": len(bags),
            "overflow_total": overflow_total,
            "commercial_pkgs": pkg_info.get("commercial"),
            "total_pkgs": total_pkgs,
            "bags_detail": bags_detail,
            "overflow_agg": overflow_agg,
            "overflow_seq": ov_seq,
            "combined": combined
        })

    # Sort: wave time group order then alpha+numeric route_short
    times_sorted = sorted({r.get("wave_time", "") for r in routes if r.get("wave_time", "")}, key=lambda x: _time_to_minutes(x) or 10**9)
    wave_rank = {t: i for i, t in enumerate(times_sorted, start=1)}

    def _route_sort_key(r: dict):
        wt = r.get("wave_time", "")
        return (wave_rank.get(wt, 999),) + _sort_route_short(r.get("route_short", ""))

    routes.sort(key=_route_sort_key)
    return routes


def build_wave_labels(routes: List[dict]) -> dict:
    times_sorted = sorted({r.get("wave_time", "") for r in routes if r.get("wave_time", "")}, key=lambda x: _time_to_minutes(x) or 10**9)
    suffix = {1: "st", 2: "nd", 3: "rd"}
    out = {}
    for i, t in enumerate(times_sorted, start=1):
        out[t] = f"{i}{suffix.get(i, 'th')} wave"
    return out


# ----------------------------- HTML template (unchanged behavior) -----------------------------
HTML_TEMPLATE = r"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1"/>
<title>__HEADER_TITLE__</title>
<style>
:root{--bg:#0b0f14;--panel:#0f1722;--text:#e8eef6;--muted:#97a7bd;--border:#1c2a3a;--accent:#3fa7ff;--page-pad-x:clamp(16px, 2.5vw, 24px);--page-pad-y:clamp(12px, 2vh, 18px);--waveColor:rgba(255,255,255,.22);}
*, *::before, *::after{box-sizing:border-box}
html,body{height:100%;width:100%}
body{margin:0;min-height:100vh;overflow-y:auto;font-family:system-ui,-apple-system,Segoe UI,Roboto,Arial,sans-serif;background:radial-gradient(1400px 800px at 20% 0%, #101826, var(--bg));color:var(--text);}
.organizerPage{width:100%;max-width:100%;min-width:0;margin:0 auto;box-sizing:border-box;padding-block:calc(var(--page-pad-y) + env(safe-area-inset-top, 0px)) calc(var(--page-pad-y) + env(safe-area-inset-bottom, 0px));padding-inline:calc(clamp(12px, 2vw, 24px) + env(safe-area-inset-left, 0px)) calc(clamp(12px, 2vw, 24px) + env(safe-area-inset-right, 0px));min-height:100vh;height:100dvh;display:flex;flex-direction:column;}
.organizerHeader{flex:0 0 auto;display:flex;flex-direction:column;gap:12px;min-width:0}
.organizerBody{flex:1 1 auto;min-height:0;width:100%;max-width:100%;overflow-y:auto;overflow-x:auto;padding:clamp(12px, 2vh, 18px) 0 0;display:flex;flex-direction:column}
.content{flex:1 1 auto;min-height:0;display:flex;flex-direction:column}
.organizerRoot{width:100%;max-width:none;min-width:0;margin:0}
.controls{display:flex;flex-direction:column;gap:12px;min-width:0;width:100%}
.header{display:flex;flex-direction:column;gap:12px;min-width:0}
.topbar{background:rgba(0,0,0,.25);border:1px solid var(--border);border-radius:14px;padding:12px 12px;min-width:0;overflow-x:auto}
.organizerHeaderRow{display:flex;align-items:center;gap:16px;flex-wrap:nowrap;min-width:0}
.topbar > *{min-width:0}
.brand{font-weight:900;white-space:nowrap;flex:0 0 auto}
.brand .routeDate{font-weight:800;}
.filterRow{display:flex;align-items:center;gap:12px;flex:1 1 360px;min-width:0}
.sel{margin-left:10px;flex:0 0 auto}
select,input{background:rgba(255,255,255,.04);border:1px solid var(--border);color:var(--text);border-radius:12px;padding:10px 12px}
select{min-width:0;width:max-content;max-width:100%;color-scheme: dark;}
/* Make native dropdown readable (Chrome/Windows) */
select option{background:#0f1722;color:#e8eef6;}
select optgroup{background:#0b0f14;color:#97a7bd;font-weight:900;}
#routeSel{
  box-shadow: inset 0 -3px 0 var(--waveColor);
  border-color: var(--waveColor);
}

.filterRow .sel{margin-left:0}
input{min-width:140px;flex:1 1 auto;width:auto}
.topbar .downloadBtn{margin-left:auto;}
.topbarBreak{display:none;}
.topCounts{
  display:flex;
  align-items:center;
  gap:14px;
  grid-column:3;
  grid-row:1;
  justify-self:end;
}
.countPill{
  opacity:.9;
  font-weight:700;
  padding:8px 12px;
  border:1px solid var(--border);
  border-radius:999px;
  --pill-bg: rgba(255,255,255,.03);
  background:var(--pill-bg);
  min-height:40px;
  display:inline-flex;
  align-items:center;
  gap:6px;
}
.progressPill{
  background:linear-gradient(90deg, var(--pill-fill) 0 var(--pill-progress, 0%), var(--pill-bg) var(--pill-progress, 0%));
}
.countPillPackages{ --pill-fill: rgba(255, 165, 79, 0.65); }
.topCountsExtra{
  display:flex;
  flex-direction:column;
  gap:8px;
  grid-column:3;
  grid-row:2;
  justify-self:end;
  align-items:flex-end;
}
.topCounts .dot{opacity:.75}
.topCounts .dot.dot-muted{background:#97a7bd}
.topCounts .countGroup{display:inline-flex;align-items:center;gap:6px}
.topCounts .countLabel{margin-left:2px}
.sectionHeaderRow{
  display:grid;
  grid-template-columns:minmax(0, 1fr) auto minmax(0, 1fr);
  grid-template-rows:auto auto;
  align-items:center;
  gap:16px;
  width:100%;
  margin-top:10px;
}
.sectionLeft{
  display:contents;
}
.sectionRight{
  display:contents;
}
.routeTitle{
  min-height:40px;
  display:flex;
  align-items:center;
  font-weight:900;
  justify-content:center;
  width:100%;
  text-align:center;
  font-size:clamp(18px, 2.2vw, 26px);
  letter-spacing:.6px;
  line-height:1.2;
  padding:4px 0;
  grid-column:2;
  grid-row:1;
  justify-self:center;
}
#routeTitle{ text-decoration: none !important; }
.sectionMeta{
  text-align:left;
  opacity:.7;
  font-size:13px;
}
.tabsRow{
  display:flex;
  gap:10px;
  margin:0 !important;
  min-height:40px;
  align-items:center;
  justify-content:flex-start;
  grid-column:1;
  grid-row:1;
  justify-self:start;
}
.tab{padding:8px 12px;border:1px solid var(--border);border-radius:999px;background:rgba(255,255,255,.03);cursor:pointer;font-weight:700;user-select:none}
.tab.active{background:rgba(255,255,255,.10)}
.card{margin-top:14px;border:1px solid var(--border);border-radius:18px;background:rgba(0,0,0,.22);padding:14px;min-width:0}
.content{margin-top:0;width:100%;max-width:none;min-width:0;display:flex;flex-direction:column}
.cardContent{
  display:flex;
  flex-direction:column;
  gap:clamp(10px, 2vh, 18px);
  padding:clamp(12px, 3vh, 24px);
  flex:1 1 auto;
  min-height:0;
}
.card.plain{background:transparent;border:none;padding:0;}
.hint{color:var(--muted);font-size:12px;margin-top:4px}
.badge{display:inline-flex;align-items:center;gap:6px;font-size:12px;color:var(--muted)}
.dot{width:8px;height:8px;border-radius:99px;background:var(--accent)}

/* tote cards */

/* tote cards */
.toteGridFrame{
  width:100%;
  min-width:0;
  display:flex;
  flex:1 1 auto;
  min-height:0;
  padding:clamp(12px, 2.4vh, 20px);
  box-sizing:border-box;
}
.toteWrap{
  width:100%;
  min-width:0;
  display:flex;
  flex:1 1 auto;
  min-height:0;
  overflow:hidden;
  box-sizing:border-box;
}
.toteBoard{
  width:100%;
  padding:0;
  min-width:0;
  display:flex;
  flex:1 1 auto;
  min-height:0;
}
.bagsGrid{
  --tote-rows:3;
  --tote-cols:1;
  --tote-scale:1;
  --tote-min-cell-w: 0px;
  --tote-gap: clamp(8px, 1.4vw, 16px);
  --tote-base-w: 210px;
  --tote-base-h: 190px;
  --tote-min-scale: 0.55;
  --tote-max-scale: 1.15;
  display:grid;
  grid-template-rows:repeat(var(--tote-rows), minmax(0, 1fr));
  grid-template-columns:repeat(var(--tote-cols), minmax(var(--tote-min-cell-w, 0px), 1fr));
  grid-auto-flow:column;
  gap:var(--tote-gap);
  justify-content:stretch;
  align-items:stretch;
  width:100%;
  height:100%;
  min-width:0;
  max-width:none;
  box-sizing:border-box;
  overflow:visible;
  flex:1 1 auto;
  direction:rtl;
}
.organizer-grid{
  overflow:visible;
}
/* grid-fit-patch: prevent horizontal scroll; allow vertical via page */
.toteWrap{
  overflow:hidden;
}
.cardsWrap, .organizer-grid{
  overflow-x:hidden;
  overflow-y:visible;
}
.toteBoard{flex:1 1 auto}
.toteCol{display:flex;flex-direction:column;gap:14px;}

.toteCard{
  --card-scale: var(--tote-scale, 1);
  --tote-badge-width: calc(44px * var(--card-scale));
  position:relative;
  width:100%;
  min-width:0;
  height:100%;
  min-height:0;
  max-width:100%;
  max-height:none;
  border-radius:18px;
  background:rgba(10,14,20,.72);
  border:1px solid rgba(255,255,255,.08);
  box-shadow: 0 10px 28px rgba(0,0,0,.35), 0 2px 0 rgba(0,0,0,.10);
  box-sizing:border-box;
  overflow:visible;
  cursor:pointer;
  container-type:inline-size;
  direction:ltr;
  display:flex;
  flex-direction:column;
  gap:10px;
  padding:6px 8px 12px;
}
.toteCard *{box-sizing:border-box;}
.toteCard.draggable{cursor:grab;}
.toteCard.dragging{opacity:.25;}
.toteCard.dropTarget{outline:2px dashed rgba(90,170,255,.85); outline-offset:2px;}
.toteCard.loaded{filter:grayscale(.85) brightness(.72);}
.toteSlot{
  border:0;
  border-radius:18px;
  background:transparent;
  display:flex;
  align-items:center;
  justify-content:center;
  color:transparent;
  font-size:12px;
}
.toteSlot.dropTarget{outline:2px dashed rgba(90,170,255,.85); outline-offset:2px;}
.customSlotControls{
  display:flex;
  flex-wrap:wrap;
  gap:8px;
  justify-content:center;
  margin-top:10px;
}
.customSlotControls .clearBtn{
  height:38px;
  padding:0 14px;
}

.toteTopRow{
  display:grid;
  grid-template-columns:minmax(var(--tote-badge-width), 1fr) minmax(0, 1.35fr) minmax(var(--tote-badge-width), 1fr);
  align-items:center;
  column-gap:6px;
  min-height:26px;
  margin:0;
}
.toteBadgeGroup{
  display:flex;
  align-items:center;
  gap:6px;
  justify-self:start;
  grid-column:1;
}
.toteRightBadge{
  display:flex;
  align-items:center;
  justify-content:flex-end;
  justify-self:end;
  grid-column:3;
  min-height:22px;
}
.toteRightBadge:empty{
  display:none;
}
.toteBubble{
  min-width:calc(22px * var(--card-scale));
  height:calc(22px * var(--card-scale));
  padding:0 6px;
  display:flex;
  align-items:center;
  justify-content:center;
  border-radius:999px;
  font-weight:900;
  font-size:calc(12px * var(--card-scale));
  line-height:1;
  text-align:center;
}
.toteCornerBadge{
  position:static;
  background:rgba(0,0,0,.72);
  border:1px solid rgba(255,255,255,.16);
  z-index:2;
  justify-self:start;
  grid-column:1;
  width:var(--tote-badge-width);
  padding:0;
}
.card-bar{
  grid-column:2;
  display:flex;
  align-items:center;
  align-self:center;
  gap:8px;
  width:calc(100% - (var(--top-gap, 0px) * 2));
  max-width:100%;
  margin:0 auto;
  justify-self:center;
  min-width:40px;
  min-height:calc(22px * var(--card-scale));
}
.bar-left-icon,
.bar-count{
  display:flex;
  align-items:center;
  justify-content:center;
  min-height:1px;
}
.bar-left-icon:empty{
  display:none;
}
.bar-track{
  flex:1 1 auto;
  position:relative;
  height:6px;
  background:rgba(255,255,255,0.22);
  border-radius:999px;
  overflow:hidden;
  box-shadow: inset 0 0 0 0.5px rgba(255,255,255,.18);
  align-self:center;
}
.bar-fill{
  height:100%;
  border-radius:999px;
  width:100%;
  background: linear-gradient(90deg, var(--chipL, #2a74ff) 0 50%, var(--chipR, var(--chipL, #2a74ff)) 50% 100%);
  border:0.5px solid color-mix(in srgb, var(--chipBorder, #000) 40%, transparent);
  box-sizing:border-box;
}
.bar-fill.yellow{
  background: linear-gradient(
    90deg,
    #FFD84D,
    #FFC400
  );
}
.bar-track,
.bar-fill{
  pointer-events:none;
}

.toteIdx{
  width:26px; height:26px;
  display:flex; align-items:center; justify-content:center;
  border-radius:999px;
  background:rgba(0,0,0,.72);
  border:1px solid rgba(255,255,255,.16);
  font-weight:900;
  font-size:calc(12px * var(--card-scale));
  flex-shrink:0;
}

.totePkg{
  background:rgba(0,0,0,.72);
  border:1px solid rgba(255,255,255,.16);
  color:#ff4b4b;
  flex-shrink:0;
  z-index:2;
  width:var(--tote-badge-width);
  padding:0;
}

.toteStar{
  width:22px; height:22px;
  display:flex; align-items:center; justify-content:center;
  border-radius:8px;
  background:rgba(0,0,0,.55);
  border:1px solid rgba(255,255,255,.10);
  font-weight:900;
  font-size:calc(16px * var(--card-scale));
  color:#ff4b4b;
  line-height:1;
  user-select:none;
  flex-shrink:0;
  position:absolute;
  z-index:2;
}
.toteStar.combine{
  color:#e3c056;
  border-color: rgba(227,192,86,.35);
  background: rgba(227,192,86,.16);
  width:auto;
  min-width:calc(22px * var(--card-scale));
  height:calc(22px * var(--card-scale));
  border-radius:999px;
  font-size:calc(12px * var(--card-scale));
  padding:0 6px;
  position:static;
}
.bags-tab .toteStar.combine{
  position:absolute;
  bottom:10px;
  left:10px;
}
.bags-tab .toteStar.on{
  bottom:10px;
  right:10px;
}
.toteStar.on{
  border-color: rgba(255,75,75,.45);
  background: rgba(255,75,75,.08);
  bottom:10px;
  right:10px;
}
.toteBigNumber{
  flex: 1 1 auto;
  min-height: 0;
  display:flex;
  align-items:center;
  justify-content:center;
  text-align:center;
  font-weight:800;
  line-height:1;
  font-size:calc(56px * var(--card-scale));
  margin:0;
  padding:2px 0;
  letter-spacing:1px;
  white-space:nowrap;
  position:static !important;
}
.toteBigNumberStack{
  flex-direction:column;
  gap: clamp(0px, 0.6cqi, 4px);
}
.toteBigNumberLine{
  font-weight:900;
  letter-spacing:1px;
  line-height:1;
  max-width:100%;
  white-space:nowrap;
  font-size:calc(48px * var(--card-scale));
}
.toteBottomRow{
  display:flex;
  flex-direction:column;
  justify-content:center;
  gap:4px;
  align-items:center;
  line-height:1.2;
  margin-top:auto;
  min-height:calc(28px * var(--card-scale));
  padding-bottom:0;
  text-align:center;
  font-weight:800;
  letter-spacing:.2px;
  opacity:.92;
  font-size:calc(13px * var(--card-scale));
  position:static !important;
  overflow:visible;
}
.toteFooter{
  white-space:nowrap;
  overflow:visible;
}
.ovZone{color:inherit;}
.ovZone99{color:#b46bff;}
.toteBottomRow .ovLine{line-height:1.25;}

/* tables */
table{width:100%;border-collapse:separate;border-spacing:0}
th,td{padding:10px 10px;border-bottom:1px solid rgba(255,255,255,.06)}

/* Overflow layout */
.ovTable{width:max-content;min-width:100%;table-layout:auto}
.ovTable td:nth-child(2){white-space:nowrap}
.ovTable td:nth-child(4){white-space:nowrap}

.ovWrap{width:100%;max-width:100%;margin:0;padding:0 18px;overflow-x:auto}
.controlsRow{min-width:0;display:flex;justify-content:space-between;align-items:center;gap:12px;flex-wrap:wrap}
.controlsRow > *{min-width:0}
.ovHeader{display:flex;justify-content:space-between;align-items:center;gap:12px;flex-wrap:wrap}
.ovHeader > *{min-width:0}
.ovHeaderRight{display:flex;align-items:center;gap:10px;flex-wrap:wrap}
.ovTitleRow{display:flex;align-items:center;gap:10px;flex-wrap:wrap}
.bagFooter{
  display:grid;
  grid-template-columns:repeat(3, minmax(0, 1fr));
  align-items:center;
  gap:12px;
  margin-top:12px;
}
.clearRow{
  display:flex;
  gap:10px;
  flex-wrap:wrap;
  align-items:center;
  justify-content:flex-end;
  justify-self:end;
}
.bagModeDock{
  display:flex;
  justify-content:flex-start;
  flex:1 1 auto;
  justify-self:start;
}
.footerCounts{
  display:flex;
  align-items:center;
  justify-content:center;
  gap:10px;
  flex-wrap:wrap;
  justify-self:center;
}
.footerCounts.singlePill .countPillCommercial{
  display:none;
}
.modeToggle{
  display:inline-flex;
  align-items:center;
  gap:0;
  background:rgba(255,255,255,.06);
  border:1px solid rgba(255,255,255,.10);
  border-radius:999px;
  overflow:hidden;
}
.modeBtn{
  padding:8px 12px;
  border:0;
  background:transparent;
  color:var(--muted);
  font-weight:800;
  cursor:pointer;
}
.modeBtn.active{background:rgba(255,255,255,.10);color:var(--fg)}
.downloadRow{flex-basis:100%;display:flex;justify-content:flex-end}
.downloadBtn{
  display:inline-flex;align-items:center;justify-content:center;
  padding:6px 12px;border-radius:999px;
  border:1px solid rgba(140,170,200,.6);
  background:#3fa7ff;color:#001018;
  font-weight:900;text-decoration:none;letter-spacing:.02em;white-space:nowrap;flex:0 0 auto;
}
.downloadLabelShort{display:none;}
.downloadBtn:hover{filter:brightness(1.08)}
.downloadBtn:active{transform:translateY(1px)}
.clearBtn{
  display:inline-flex;
  align-items:center;
  justify-content:center;
  padding:6px 12px;
  border-radius:999px;
  border:1px solid rgba(255,255,255,.16);
  background:rgba(255,255,255,.08);
  color:var(--fg);
  font-weight:900;
  cursor:pointer;
}
.clearBtn:hover{filter:brightness(1.08)}
.clearBtn:active{transform:translateY(1px)}
.syncBtn{
  padding:6px 12px;
  border-radius:999px;
  border:1px solid rgba(140,170,200,.6);
  background:#5f7fa6;
  color:#0b0f14;
  font-weight:900;
  cursor:pointer;
}
.syncBtn:hover{filter:brightness(1.08)}
.syncBtn:active{transform:translateY(1px)}

/* Overflow checklist cells */
.ovChecks{display:flex;flex-wrap:wrap;gap:8px;align-items:center;justify-content:flex-start;min-height:22px}
.ovBox{
  width:18px;height:18px;border-radius:5px;
  border:1px solid rgba(255,255,255,.22);
  background:rgba(0,0,0,.25);
  box-shadow: inset 0 0 0 1px rgba(0,0,0,.25);
  cursor:pointer;
}
.ovBox.on{
  background:rgba(255,255,255,.22);
  border-color:rgba(255,255,255,.38);
}
tr.ovDone td{
  color:#aab6c6;
  background:rgba(255,255,255,.10);
}
tr.ovDone{box-shadow: inset 0 0 0 2px rgba(255,255,255,.10);}
tr.ovDone td{filter: grayscale(1);}
tr.ovDone td .ovBox{opacity:.85}
.ovLoadedPill{margin-left:10px;display:inline-flex;align-items:center;gap:6px;padding:4px 10px;border-radius:999px;font-weight:900;font-size:12px;letter-spacing:.02em;background:rgba(255,255,255,.14);border:1px solid rgba(255,255,255,.14);color:var(--fg)}
.ovMode{
  display:inline-flex;align-items:center;gap:0;
  background:rgba(255,255,255,.06);
  border:1px solid rgba(255,255,255,.10);
  border-radius:999px;
  overflow:hidden;
}
.ovMode button{
  padding:8px 12px;border:0;background:transparent;color:var(--muted);font-weight:800;cursor:pointer;
}
.ovMode button.on{background:rgba(255,255,255,.10);color:var(--fg)}
tr.ovDrag{cursor:grab}
tr.ovDrag.dragging{opacity:.55}
tr.ovDrag.dropTarget{outline:2px dashed rgba(255,255,255,.25); outline-offset:-6px}
th{color:var(--muted);font-size:12px;text-align:left}
td:last-child,th:last-child{text-align:right}

@media (max-width: 1100px){
  .organizerPage{padding:16px;min-height:100vh;}
}

@media (max-width: 720px){
  .brand .routeCode,
  .brand .routeSep{display:none;}
  .downloadLabelFull{display:none;}
  .downloadLabelShort{display:inline;}
  .downloadBtn{padding:6px 10px;font-size:12px;}
  .sectionHeaderRow{
    display:flex;
    flex-wrap:nowrap;
    align-items:center;
    gap:8px;
  }
  .routeTitle{
    flex:1 1 auto;
    min-width:0;
    justify-content:flex-start;
    text-align:left;
    font-size:clamp(14px, 4vw, 18px);
    white-space:nowrap;
    overflow:hidden;
    text-overflow:ellipsis;
  }
  .topCounts{
    flex:0 0 auto;
    gap:6px;
  }
  .tabsRow{
    flex:0 0 auto;
    gap:6px;
  }
  .countPill{padding:6px 8px;min-height:32px;font-size:12px;}
  .tab{padding:4px 8px;font-size:11px;white-space:nowrap;}
}

@media (max-width: 720px) and (orientation: portrait){
  .organizerHeaderRow{flex-wrap:wrap;align-items:center;gap:10px;row-gap:0;column-gap:10px;}
  .topbar .brand{order:1;}
  .topbar .downloadBtn{order:2;}
  .topbarBreak{display:block;flex-basis:100%;height:0;order:3;}
  .topbar .filterRow{order:4;flex:1 1 100%;min-width:0;}
  .topbar .sel{flex:0 1 45%;min-width:0;}
  .topbar .sel select{width:100%;}
  .topbar #q{flex:1 1 55%;min-width:0;font-size:12px;}
}

/* FULL-WIDTH OVERRIDE */
.organizerPage,
.organizerRoot{
  width:100% !important;
  max-width:100% !important;
  margin:0 auto !important;
  box-sizing:border-box !important;
}
.organizerPage,
.organizerHeader,
.panel,
.shell,
.container{
  overflow:visible !important;
}

/* Combined tab is shown by default */

</style>
</head>
<body>
<div class="organizerRoot organizerPage">
  <div class="organizerHeader">
    <div class="controls">
      <div class="header">
        <div class="topbar organizerHeaderRow">
          <div class="brand">
            <span class="routeCode">__HEADER_ROUTE_CODE__</span>
            <span class="routeSep">__HEADER_ROUTE_SEP__</span>
            <span class="routeDate">__HEADER_ROUTE_DATE__</span>
          </div>
          <div class="filterRow">
            <div class="sel"><select id="routeSel"></select></div>
            <input id="q" placeholder="Search Bag / Overflow Info"/>
          </div>
          <a class="downloadBtn" href="download/STACKED.pdf">
            <span class="downloadLabelFull">PDF</span>
            <span class="downloadLabelShort">PDF</span>
          </a>
          <span class="topbarBreak" aria-hidden="true"></span>
        </div>

        <div class="sectionHeaderRow">
          <div class="sectionLeft">
            <div class="routeTitle" id="routeTitle"></div>
          </div>
          <div class="sectionRight">
            <div class="topCounts countPill">
              <span class="countGroup">
                <span class="dot"></span>
                <span id="bagsCount">0</span>
                <span class="countLabel">bags</span>
              </span>
              <span class="countGroup">
                <span class="dot dot-muted"></span>
                <span id="ovCount">0</span>
                <span class="countLabel">overflow</span>
              </span>
            </div>
            <div class="tabsRow">
              <div class="tab active" data-tab="combined">Bags + Overflow</div>
              <div class="tab" data-tab="bags">Bags</div>
              <div class="tab" data-tab="overflow">Overflow</div>
            </div>
          </div>
        </div>
      </div>
    </div>
  </div>

  <div class="organizerBody">
    <div id="content" class="card content cardContent"></div>
  </div>
</div>

<script>
const ROUTES = __ROUTES_JSON__;
const WAVE_LABEL_BY_TIME = __WAVE_JSON__;
const organizerRoot = document.querySelector(".organizerRoot");
const organizerBody = document.querySelector(".organizerBody");
const selectMeasureCanvas = document.createElement("canvas");

let renderRaf = 0;

function scheduleRender(){
  if(renderRaf) return;
  renderRaf = requestAnimationFrame(()=>{
    renderRaf = 0;
    render();
  });
}

// --- Persistent state ---
const STORAGE_KEY = "vanorg_loaded_v1";
const MODE_KEY = "vanorg_bagmode_v1";
const ORDER_KEY = "vanorg_bagorder_v1";
const COMBINE_KEY = "vanorg_combined_v1";
const CUSTOM_SLOTS_KEY = "vanorg_custom_slots_v1";
const LAST_MODE_KEY = "vanorg_last_bagmode_v1";

function readJSON(key, fallback){ try { return JSON.parse(localStorage.getItem(key) || JSON.stringify(fallback)); } catch(e){ return fallback; } }
function writeJSON(key, obj){ try { localStorage.setItem(key, JSON.stringify(obj)); } catch(e){} }

let LOADED = readJSON(STORAGE_KEY, {});
let BAGMODE = readJSON(MODE_KEY, {});
let BAGORDER = readJSON(ORDER_KEY, {});
let COMBINED = readJSON(COMBINE_KEY, {});
let CUSTOMSLOTS = readJSON(CUSTOM_SLOTS_KEY, {});
let LAST_NON_CUSTOM = readJSON(LAST_MODE_KEY, {});

// Overflow checklist + ordering
const OVKEY = "vanorg_overflow_checks_v1";
const OVMODE_KEY = "vanorg_overflow_mode_v1";
const OVORDER_KEY = "vanorg_overflow_order_v1";

let OVCHK = readJSON(OVKEY, {});
let OVMODE = readJSON(OVMODE_KEY, {});
let OVORDER = readJSON(OVORDER_KEY, {});
let RESET_ARMED = {};

function clearResetArmed(routeShort){
  RESET_ARMED[routeShort] = false;
}

function getOvMode(routeShort){ return OVMODE[routeShort] || "normal"; }
function setOvMode(routeShort, mode){ OVMODE[routeShort] = mode; writeJSON(OVMODE_KEY, OVMODE); }

function getOvOrder(routeShort){ return (OVORDER[routeShort] || []).slice(); }
function setOvOrder(routeShort, arr){ OVORDER[routeShort] = arr.slice(); writeJSON(OVORDER_KEY, OVORDER); }

function isOvChecked(routeShort, rowId, k){
  return !!(OVCHK[routeShort] && OVCHK[routeShort][rowId] && OVCHK[routeShort][rowId][String(k)]);
}
function toggleOvChecked(routeShort, rowId, k){
  if(!OVCHK[routeShort]) OVCHK[routeShort] = {};
  if(!OVCHK[routeShort][rowId]) OVCHK[routeShort][rowId] = {};
  const kk = String(k);
  if(OVCHK[routeShort][rowId][kk]) delete OVCHK[routeShort][rowId][kk];
  else OVCHK[routeShort][rowId][kk] = true;
  writeJSON(OVKEY, OVCHK);
}
function clearOvRoute(routeShort){
  OVCHK[routeShort] = {};
  OVMODE[routeShort] = "normal";
  OVORDER[routeShort] = [];
  writeJSON(OVKEY, OVCHK);
  writeJSON(OVMODE_KEY, OVMODE);
  writeJSON(OVORDER_KEY, OVORDER);
}

function isLoaded(routeShort, idx){ return !!(LOADED[routeShort] && LOADED[routeShort][String(idx)]); }
function toggleLoaded(routeShort, idx){
  if(!LOADED[routeShort]) LOADED[routeShort] = {};
  const k = String(idx);
  if(LOADED[routeShort][k]) delete LOADED[routeShort][k];
  else LOADED[routeShort][k] = true;
  writeJSON(STORAGE_KEY, LOADED);
  clearResetArmed(routeShort);
}
function clearLoaded(routeShort){ LOADED[routeShort] = {}; writeJSON(STORAGE_KEY, LOADED); }

function getMode(routeShort){ return BAGMODE[routeShort] || "normal"; }
function setMode(routeShort, mode){
  BAGMODE[routeShort] = mode;
  writeJSON(MODE_KEY, BAGMODE);
  if(mode !== "custom"){
    setLastNonCustomMode(routeShort, mode);
  }
}

function getLastNonCustomMode(routeShort){
  const stored = LAST_NON_CUSTOM[routeShort];
  if(stored === "normal" || stored === "reversed") return stored;
  const current = getMode(routeShort);
  return current === "reversed" ? "reversed" : "normal";
}
function setLastNonCustomMode(routeShort, mode){
  if(mode !== "normal" && mode !== "reversed") return;
  LAST_NON_CUSTOM[routeShort] = mode;
  writeJSON(LAST_MODE_KEY, LAST_NON_CUSTOM);
}

function getCustomSlots(routeShort){ return (CUSTOMSLOTS[routeShort] || []).slice(); }
function setCustomSlots(routeShort, slots){ CUSTOMSLOTS[routeShort] = slots.slice(); writeJSON(CUSTOM_SLOTS_KEY, CUSTOMSLOTS); }

function isCombinedSecond(routeShort, secondIdx){ return !!(COMBINED[routeShort] && COMBINED[routeShort][String(secondIdx)]); }
function setCombined(routeShort, secondIdx, val){
  if(!COMBINED[routeShort]) COMBINED[routeShort] = {};
  const k = String(secondIdx);
  if(val) COMBINED[routeShort][k] = true;
  else delete COMBINED[routeShort][k];
  writeJSON(COMBINE_KEY, COMBINED);
  clearResetArmed(routeShort);
}
function clearCombined(routeShort){
  COMBINED[routeShort] = {};
  writeJSON(COMBINE_KEY, COMBINED);
}
function resetBagsPage(routeShort, baseOrderArr, items){
  // Unpress all totes + uncombine everything
  clearLoaded(routeShort);
  clearCombined(routeShort);
  // If in custom mode, restore order to base (so nothing stays missing)
  if(getMode(routeShort)==="custom"){
    setCustomOrder(routeShort, baseOrderArr.slice());
    if(items){
      setCustomSlots(routeShort, defaultCustomSlots(items));
    }
  }
}


let activeRouteIndex = 0;
let activeTab = "combined";


const routeSel = document.getElementById("routeSel");
const qBox = document.getElementById("q");
const content = document.getElementById("content");

let WAVE_COLORS = {}; // { "HH:MM": "#RRGGBB" }

function timeKey(label){
  const m = String(label||"").match(/(\d{1,2}):(\d{2})\s*([AP]M)?/i);
  if(!m) return "";
  let hh = parseInt(m[1],10);
  const mm = m[2];
  const ap = (m[3]||"").toUpperCase();
  if(ap==="PM" && hh!==12) hh += 12;
  if(ap==="AM" && hh===12) hh = 0;
  return String(hh).padStart(2,"0")+":"+mm;
}

function hexToRgb(hex){
  const h = (hex||"").replace("#","").trim();
  if(h.length!==6) return null;
  return { r:parseInt(h.slice(0,2),16), g:parseInt(h.slice(2,4),16), b:parseInt(h.slice(4,6),16) };
}

function rgbToHue(r,g,b){
  r/=255; g/=255; b/=255;
  const max=Math.max(r,g,b), min=Math.min(r,g,b);
  const d=max-min;
  if(d===0) return 0;
  let h;
  if(max===r) h=((g-b)/d)%6;
  else if(max===g) h=(b-r)/d+2;
  else h=(r-g)/d+4;
  h=Math.round(h*60);
  if(h<0) h+=360;
  return h;
}

function waveEmoji(hex){
  const rgb = hexToRgb(hex);
  if(!rgb) return "⚪️";
  const {r,g,b}=rgb;
  const avg=(r+g+b)/3;
  if(avg < 55) return "⚫️";
  if(avg > 235) return "⚪️";
  const h = rgbToHue(r,g,b);
  if(h < 20 || h >= 340) return "🔴";
  if(h < 50) return "🟠";
  if(h < 75) return "🟡";
  if(h < 165) return "🟢";
  if(h < 255) return "🔵";
  return "🟣";
}

function waveColorForRoute(r){
  const key = timeKey(r.wave_time||"");
  return (key && WAVE_COLORS[key]) ? WAVE_COLORS[key] : "";
}

function applyWaveUI(r){
  const c = waveColorForRoute(r);
  document.documentElement.style.setProperty("--waveColor", c || "rgba(255,255,255,.22)");
}

function rebuildDropdownWithWaveDots(){
  if(!routeSel) return;
  routeSel.querySelectorAll("option").forEach(opt=>{
    const idx = parseInt(opt.value,10);
    const r = ROUTES[idx];
    if(!r) return;
    const c = waveColorForRoute(r);
    const dot = c ? waveEmoji(c) : "⚪️";
    opt.textContent = `${dot} ${routeTitle(r)}`;
  });
  adjustRouteSelectWidth();
}

function updateSearchPlaceholder(){
  if(!qBox) return;
  const portraitMatch = window.matchMedia("(orientation: portrait) and (max-width: 720px)").matches;
  qBox.placeholder = portraitMatch ? "Search Bag / Overflow" : "Search Bag / Overflow Info";
}

updateSearchPlaceholder();
window.addEventListener("resize", updateSearchPlaceholder);
window.addEventListener("orientationchange", updateSearchPlaceholder);

function routeTitle(r){ return (r.route_short||"") + (r.cx ? ` (${r.cx})` : ""); }
function baseOrder(r){ return (r.bags_detail||[]).map(x=>x.idx); }

function getFooterPackagePillWidth(){
  const pill = document.querySelector("#footerCounts .countPill:not(.countPillCommercial)");
  if(!pill) return null;
  const rect = pill.getBoundingClientRect();
  if(!rect || !rect.width) return null;
  return Math.round(rect.width);
}

function sendMetaToParent(r){
  try{
    const stats = getLoadedStats(r);
    const footerWidth = getFooterPackagePillWidth();
    window.parent.postMessage({
      type: "routeMeta",
      title: routeTitle(r),
      bags: r.bags_count ?? 0,
      bags_loaded: stats.loadedCards,
      overflow: r.overflow_total ?? 0,
      overflow_loaded: stats.overflowLoaded,
      commercial: r.commercial_pkgs ?? null,
      total: r.total_pkgs ?? null,
      total_loaded: stats.totalLoaded,
      footer_pill_width: footerWidth
    }, "*");
  }catch(e){}
}

function sumOverflowCountsForBag(entry, ovMap){
  if(!entry || !ovMap) return 0;
  const label = String(entry.bag || entry.bag_id || "").trim();
  if(!label) return 0;
  const list = ovMap.get(bagKey(label));
  if(!list || !list.length) return 0;
  return list.reduce((acc, item)=>acc + (parseInt(item.count || 0, 10) || 0), 0);
}

function pkgFooterCounts(bag){
  if(!bag) return { base: null, overflow: 0 };
  const val = bag.pkgs;
  if(val === undefined || val === null || val === "") return { base: null, overflow: 0 };
  const str = String(val);
  const baseMatch = str.match(/^\s*(-?\d+)/);
  const base = baseMatch ? parseInt(baseMatch[1], 10) : null;
  const overflowMatch = str.match(/\((\d+)\)/);
  const overflow = overflowMatch ? parseInt(overflowMatch[1], 10) || 0 : 0;
  return { base, overflow };
}

function pkgOverflowValue(bag){
  return pkgFooterCounts(bag).overflow;
}

function overflowCountForEntry(entry, ovMap){
  if(!entry) return 0;
  const fromMap = sumOverflowCountsForBag(entry, ovMap);
  const fromFooter = pkgOverflowValue(entry);
  return Math.max(fromMap, fromFooter);
}

function pkgCountNumber(anchor, other){
  const first = pkgCountValue(anchor);
  const second = pkgCountValue(other);
  if(first === null && second === null) return 0;
  return (first || 0) + (second || 0);
}

function getLoadedStats(r){
  const routeShort = r.route_short || "";
  const loadedEntries = routeShort && LOADED[routeShort] ? Object.keys(LOADED[routeShort]) : [];
  const byIdx = Object.fromEntries((r.bags_detail || []).map(x=>[x.idx, x]));
  const ovMap = buildOverflowMap(r);
  let overflowLoaded = 0;
  let pkgLoaded = 0;
  let loadedCards = 0;
  loadedEntries.forEach((key)=>{
    const idx = parseInt(key, 10);
    if(!idx) return;
    const isSecond = isCombinedSecond(routeShort, idx);
    if(isSecond && isLoaded(routeShort, idx - 1)) return;
    const cur = byIdx[idx];
    if(!cur && !isSecond) return;
    if(isSecond){
      const first = byIdx[idx - 1];
      if(!first && !cur) return;
      overflowLoaded += overflowCountForEntry(first, ovMap);
      overflowLoaded += overflowCountForEntry(cur, ovMap);
      pkgLoaded += pkgCountNumber(first, cur);
      loadedCards += 2;
      return;
    }
    const secondIdx = idx + 1;
    const second = isCombinedSecond(routeShort, secondIdx) ? byIdx[secondIdx] : null;
    overflowLoaded += overflowCountForEntry(cur, ovMap);
    if(second) overflowLoaded += overflowCountForEntry(second, ovMap);
    pkgLoaded += pkgCountNumber(cur, second);
    loadedCards += second ? 2 : 1;
  });
  return {
    loadedCards,
    overflowLoaded,
    pkgLoaded,
    totalLoaded: pkgLoaded + overflowLoaded
  };
}

function updateFooterCounts(r){
  const wrap = document.getElementById("footerCounts");
  const commercial = document.getElementById("commercialCount");
  const total = document.getElementById("totalCount");
  const totalLabel = document.getElementById("totalLabel");
  const packagePill = wrap ? wrap.querySelector(".countPillPackages") : null;
  if(!wrap || !commercial || !total) return;
  const routeShort = r.route_short || r.short || "";
  const loadedEntries = routeShort && LOADED[routeShort] ? Object.keys(LOADED[routeShort]) : [];
  const hasLoaded = loadedEntries && loadedEntries.length > 0;
  const hasCommercial = r.commercial_pkgs !== undefined && r.commercial_pkgs !== null;
  const hasTotal = r.total_pkgs !== undefined && r.total_pkgs !== null;
  if(!hasCommercial && !hasTotal){
    wrap.style.display = "none";
    return;
  }
  wrap.style.display = "";
  wrap.classList.toggle("singlePill", hasLoaded);
  commercial.textContent = hasCommercial ? r.commercial_pkgs : "—";
  if(hasTotal){
    const stats = getLoadedStats(r);
    const totalNum = parseInt(r.total_pkgs, 10);
    const loadedNum = parseInt(stats.totalLoaded || 0, 10);
    if(Number.isNaN(totalNum)){
      total.textContent = "—";
      if(totalLabel) totalLabel.textContent = "packages";
      if(packagePill) packagePill.style.setProperty("--pill-progress", "0%");
      return;
    }
    if(hasLoaded){
      const remaining = Math.max(totalNum - loadedNum, 0);
      total.textContent = `${loadedNum}/${totalNum}`;
      if(totalLabel) totalLabel.textContent = `packages (${remaining} left)`;
      if(packagePill){
        const pct = totalNum > 0 ? Math.max(0, Math.min(loadedNum / totalNum, 1)) : 0;
        packagePill.style.setProperty("--pill-progress", `${(pct * 100).toFixed(1)}%`);
      }
      return;
    }
    total.textContent = `${totalNum}`;
    if(totalLabel) totalLabel.textContent = "packages";
    if(packagePill) packagePill.style.setProperty("--pill-progress", "0%");
  }else{
    total.textContent = "—";
    if(totalLabel) totalLabel.textContent = "packages";
    if(packagePill) packagePill.style.setProperty("--pill-progress", "0%");
  }
}

function setActiveTab(name){
  if(!name) return;
  activeTab = name;
  render();
}

function normalizeRouteToken(val){
  return String(val || "").trim().toLowerCase();
}

function findRouteIndexFromParam(param){
  const needle = normalizeRouteToken(param);
  if(!needle) return null;
  for(let i = 0; i < ROUTES.length; i++){
    if(normalizeRouteToken(ROUTES[i].route_short) === needle) return i;
  }
  for(let i = 0; i < ROUTES.length; i++){
    if(normalizeRouteToken(routeTitle(ROUTES[i])) === needle) return i;
  }
  for(let i = 0; i < ROUTES.length; i++){
    const shortToken = normalizeRouteToken(ROUTES[i].route_short);
    if(shortToken && needle.includes(shortToken)) return i;
  }
  return null;
}

const routeParam = new URLSearchParams(window.location.search).get("route");
const routeParamIndex = findRouteIndexFromParam(routeParam);
if(routeParamIndex !== null){
  activeRouteIndex = routeParamIndex;
}

function bagModeHtml(routeShort){
  const mode = getMode(routeShort);
  return `
    <div class="modeToggle" role="tablist" aria-label="Bag order mode">
      <button class="modeBtn ${mode==="normal" ? "active":""}" data-bagmode="normal">Normal</button>
      <button class="modeBtn ${mode==="reversed" ? "active":""}" data-bagmode="reversed">Reversed</button>
      <button class="modeBtn ${mode==="custom" ? "active":""}" data-bagmode="custom">Custom</button>
    </div>
  `;
}

// Saturated chips (high-contrast)
function bagColorChip(label){
  const s = (label||"").toLowerCase();
  if (s.includes("yellow")) return "#FFD400";
  if (s.includes("orange")) return "#FF6A00";
  if (s.includes("green"))  return "#00D26A";
  if (s.includes("navy"))   return "#1E5BFF";
  if (s.includes("black"))  return "#0B0B0B";
  return "#34B3FF";
}
function chipBorderColor(chip1, chip2){
  const isBlack = (chip)=> (chip || "").toLowerCase() === "#0b0b0b";
  return (isBlack(chip1) || isBlack(chip2)) ? "#FFFFFF" : "#000000";
}

function normZone(z){
  if(!z) return "";
  let s = String(z).trim();
  s = s.replace(/^[A-Za-z]-/, "");
  const m = s.match(/(\d+\.\d+[A-Za-z])/);
  return m ? m[1] : s;
}
function parseZoneCounts(zones){
  if(!zones) return [];
  return String(zones)
    .split(';')
    .map(p=>p.trim())
    .filter(Boolean)
    .map((p)=>{
      const cnt = p.match(/(\d+\.\d+[A-Za-z])\s*\((\d+)\)/);
      if(cnt) return { zone: cnt[1], count: parseInt(cnt[2],10) || 0 };
      const zon = p.match(/(\d+\.\d+[A-Za-z])/);
      if(zon) return { zone: zon[1], count: 0 };
      return { zone: p, count: 0 };
    });
}
function match(text,q){
  if(!q) return true;
  const tokens = String(q).toLowerCase().split(/\s+/).filter(Boolean);
  if(!tokens.length) return true;
  const hay = (text||"").toLowerCase();
  return tokens.every(t=>hay.includes(t));
}

function bagLabel(entry){
  if(!entry) return "";
  return String(entry.bag_id || entry.bag || "").trim();
}

function bagKey(label){
  return String(label || "")
    .trim()
    .toUpperCase()
    .replace(/\s+/g, "");
}

function overflowZonesText(label, ovMap){
  if(!label || !ovMap) return "";
  const entry = ovMap.get(bagKey(label));
  if(!entry || !entry.length) return "";
  return entry.map((item)=>`${item.zone||""} ${normZone(item.zone)}`).join(" ");
}
function overflowSearchText(label, ovMap){
  if(!ovMap) return "";

  let v = null;
  const key = bagKey(label);

  // ovMap is a Map in this codebase
  if(ovMap instanceof Map){
    v = ovMap.get(key);
  }else{
    v = ovMap[key];
  }

  if(!v) return "";

  // Most cases are arrays like [{zone:"16.2U",count:11}, {zone:"99.6X",count:1}] or strings.
  if(Array.isArray(v)){
    return v.map((item)=>{
      if(item && typeof item === "object"){
        const zone = item.zone || "";
        const norm = normZone(zone);
        const count = item.count ? `(${item.count})` : "";
        return `${zone} ${norm} ${count}`.trim();
      }
      return String(item || "");
    }).join(" ").toLowerCase();
  }

  return String(v).toLowerCase();
}

// Custom order helpers
function slotIdForItem(item){
  if(!item) return "";
  const idx = item.idx ?? (item.cur && item.cur.idx);
  if(idx === undefined || idx === null) return "";
  return String(idx);
}

function defaultCustomSlots(items){
  const slots = (items || []).map(slotIdForItem).filter(Boolean);
  while(slots.length % 3 !== 0) slots.push(null);
  return slots;
}

function customSlotsFromOrder(orderArr){
  const slots = (orderArr || []).map((idx)=>String(idx));
  while(slots.length % 3 !== 0) slots.push(null);
  return slots;
}

function normalizeCustomSlots(routeShort, items){
  const validIds = new Set((items || []).map(slotIdForItem).filter(Boolean));
  const seen = new Set();
  const raw = getCustomSlots(routeShort);
  const normalized = raw.map((val)=>{
    if(val === null || val === undefined) return null;
    const key = String(val).trim();
    if(!key || !validIds.has(key) || seen.has(key)) return null;
    seen.add(key);
    return key;
  });
  (items || []).forEach((item)=>{
    const key = slotIdForItem(item);
    if(!key || seen.has(key)) return;
    normalized.push(key);
    seen.add(key);
  });
  while(normalized.length % 3 !== 0) normalized.push(null);
  setCustomSlots(routeShort, normalized);
  return normalized;
}

function customOrderFromSlots(routeShort, baseOrderArr){
  const baseSet = new Set(baseOrderArr.map(String));
  const seen = new Set();
  const order = [];
  getCustomSlots(routeShort).forEach((slot)=>{
    if(slot === null || slot === undefined) return;
    const key = String(slot);
    if(!baseSet.has(key) || seen.has(key)) return;
    seen.add(key);
    order.push(parseInt(key, 10));
  });
  baseOrderArr.forEach((idx)=>{
    const key = String(idx);
    if(seen.has(key)) return;
    seen.add(key);
    order.push(idx);
  });
  return order;
}

function getCustomOrder(routeShort, base){
  const arr = (BAGORDER[routeShort] || []).slice();
  const baseSet = new Set(base.map(String));
  const cleaned = arr.map(String).filter(x=>baseSet.has(x));
  const cleanedSet = new Set(cleaned);
  base.forEach(i=>{ const s=String(i); if(!cleanedSet.has(s)) cleaned.push(s); });
  BAGORDER[routeShort] = cleaned.map(x=>parseInt(x,10));
  writeJSON(ORDER_KEY, BAGORDER);
  return BAGORDER[routeShort];
}
function setCustomOrder(routeShort, orderArr){ BAGORDER[routeShort] = orderArr.map(x=>parseInt(x,10)); writeJSON(ORDER_KEY, BAGORDER); }

function removeCombinedSecondsFromOrder(routeShort, orderArr){
  const sec = COMBINED[routeShort] || {};
  const secSet = new Set(Object.keys(sec).map(x=>parseInt(x,10)));
  if(secSet.size===0) return orderArr;
  return orderArr.filter(i=>!secSet.has(i));
}

function buildOrderForMode(r, mode){
  const routeShort = r.route_short;
  const base = baseOrder(r);
  let ord = base.slice();
  if(mode==="reversed") ord.reverse();
  if(mode==="custom") ord = customOrderFromSlots(routeShort, base);
  if(mode==="custom") ord = removeCombinedSecondsFromOrder(routeShort, ord);
  return ord;
}

function buildOrder(r){
  const mode = getMode(r.route_short);
  return buildOrderForMode(r, mode);
}

function buildDisplayItems(r, q, ovMap){
  const routeShort = r.route_short;
  const byIdx = Object.fromEntries((r.bags_detail||[]).map(x=>[x.idx, x]));
  const ord = buildOrder(r);
  const items = [];
  for(const idx of ord){
    if(isCombinedSecond(routeShort, idx)) continue;
    const cur = byIdx[idx];
    if(!cur) continue;
    const secondIdx = idx + 1;
    const second = isCombinedSecond(routeShort, secondIdx) ? byIdx[secondIdx] : null;
    const eligibleCombine = (!cur.sort_zone) && idx > 1;
    // IMPORTANT: combined cards use bag_id as the tote/bag key; label/bag may be missing
    const curLabel = cur.bag_id || cur.label || cur.bag;
    const secondLabel = second && (second.bag_id || second.label || second.bag);
    const curOverflow = overflowSearchText(cur.bag || curLabel, ovMap);
    const secondOverflow = overflowSearchText((second && second.bag) || secondLabel, ovMap);
    const curSort = normZone(cur.sort_zone);
    const secondSort = second ? normZone(second.sort_zone) : "";
    const text = `${cur.idx} ${curLabel} ${cur.bag||""} ${cur.sort_zone||""} ${curSort} ${cur.pkgs||""} ${curOverflow}` +
      (second ? ` ${secondLabel} ${second.bag||""} ${second.sort_zone||""} ${secondSort} ${second.pkgs||""} ${secondOverflow}` : "");
    if(!match(text, q)) continue;
    items.push({ idx, cur, secondIdx: second ? secondIdx : null, second, eligibleCombine });
  }
  return items;
}

function pkgCountValue(bag){
  const counts = pkgFooterCounts(bag);
  if(counts.base === null || Number.isNaN(counts.base)) return null;
  return counts.base;
}

function combinedPkgSum(anchor, other){
  const first = pkgCountValue(anchor);
  const second = pkgCountValue(other);
  if(first === null && second === null) return "";
  return String((first || 0) + (second || 0));
}

function buildToteCardHtml(it, routeShort, getSubLine, getBadgeText, getPkgCount, slotIndex){
  const cur = it.cur;
  const second = it.second;
  const main1 = (cur.bag_id || cur.bag || "").toString();
  const chip1 = bagColorChip(cur.bag);
  const loadedClass = isLoaded(routeShort, it.idx) ? "loaded" : "";
  const badgeText = getBadgeText ? getBadgeText(cur, second, it.idx) : it.idx;
  const pkgText = getPkgCount ? getPkgCount(cur, second) : "";
  const badgeHtml = badgeText
    ? `<div class="toteCornerBadge toteBubble">${badgeText}</div>`
    : ``;
  const pkgHtml = pkgText ? `<div class="totePkg toteBubble">${pkgText}</div>` : ``;
  const pkgClass = pkgText ? "hasPkg" : "";
  const sortZoneClass = cur.sort_zone ? "" : "noSortZone";
  const starHtml = it.eligibleCombine ? `<div class="toteStar combine toteBubble" data-action="combine" data-second="${it.idx}" title="Combine with previous">+</div>` : ``;
  const badgeGroupHtml = `<div class="toteBadgeGroup">${badgeHtml}</div>`;
  const barHtml = `<div class="card-bar">
    <span class="bar-left-icon">${starHtml}</span>
    <div class="bar-track"><div class="bar-fill"></div></div>
  </div>`;
  const rightBadgeHtml = `<div class="toteRightBadge">${pkgHtml}</div>`;
  const slotAttr = (slotIndex === 0 || slotIndex) ? ` data-slot="${slotIndex}"` : "";
  if(second){
    const main2 = (second.bag_id || second.bag || "").toString();
    const chip2 = bagColorChip(second.bag);
    const sub = getSubLine(cur, second);
    const topNum = (cur.sort_zone ? main1 : main2);
    const botNum = (cur.sort_zone ? main2 : main1);
    const minusHtml = `<div class="toteStar on" data-action="uncombine" data-second="${it.secondIdx}" title="Uncombine">-</div>`;
    const chipBorder = chipBorderColor(chip1, chip2);
    return `<div class="toteCard ${loadedClass} ${pkgClass} ${sortZoneClass}" data-idx="${it.idx}"${slotAttr} style="--chipL:${chip1};--chipR:${chip2};--chipBorder:${chipBorder};">
      ${minusHtml}
      <div class="toteTopRow">
        ${badgeGroupHtml}
        ${barHtml}
        ${rightBadgeHtml}
      </div>
      <div class="toteBigNumber toteBigNumberStack">
        <div class="toteBigNumberLine">${topNum}</div>
        <div class="toteBigNumberLine">${botNum}</div>
      </div>
      ${sub ? `<div class="toteBottomRow toteFooter">${sub}</div>` : ``}
    </div>`;
  }

  const sub = getSubLine(cur, null);
  const chipBorder = chipBorderColor(chip1, chip1);
  return `<div class="toteCard ${loadedClass} ${pkgClass} ${sortZoneClass}" data-idx="${it.idx}"${slotAttr} style="--chipL:${chip1};--chipR:${chip1};--chipBorder:${chipBorder};">
    <div class="toteTopRow">
      ${badgeGroupHtml}
      ${barHtml}
      ${rightBadgeHtml}
    </div>
    <div class="toteBigNumber">${main1}</div>
    ${sub ? `<div class="toteBottomRow toteFooter">${sub}</div>` : ``}
  </div>`;
}

function buildToteLayout(items, routeShort, getSubLine, getBadgeText, getPkgCount){
  const orderedItems = items.slice();
  const cardsHtml = orderedItems.map((it)=>{
    return buildToteCardHtml(it, routeShort, getSubLine, getBadgeText, getPkgCount, null);
  }).join("");

  return { cardsHtml };
}

function buildCustomSlotsLayout(routeShort, slots, itemsById, getSubLine, getBadgeText, getPkgCount){
  const cardsHtml = slots.map((slot, index)=>{
    if(!slot || !itemsById.has(slot)){
      return `<div class="toteSlot" data-slot="${index}" aria-label="Empty slot"></div>`;
    }
    const item = itemsById.get(slot);
    return buildToteCardHtml(item, routeShort, getSubLine, getBadgeText, getPkgCount, index);
  }).join("");
  return { cardsHtml };
}

function buildOverflowMap(r){
  const map = new Map();
  (r.combined || []).forEach((x)=>{
    const bag = bagKey(x.bag);
    if(!bag) return;
    let entry = map.get(bag);
    if(!entry){
      entry = [];
      map.set(bag, entry);
    }
    parseZoneCounts(x.zones || "").forEach(z=>entry.push(z));
  });
  return map;
}

function overflowSummary(bagLabel, ovMap){
  if(!bagLabel || !ovMap) return "";
  const entry = ovMap.get(bagKey(bagLabel));
  if(!entry || !entry.length) return "";
  return entry.map((item)=>{
    const label = normZone(item.zone);
    if(!label) return "";
    const count = item.count ? ` (${item.count})` : "";
    const cls = label.startsWith("99.") ? "ovZone ovZone99" : "ovZone";
    return `<div class="ovLine"><span class="${cls}">${label}${count}</span></div>`;
  }).filter(Boolean).join("");
}


function attachBagHandlers(routeShort, allowDrag, customState){
  const hasCustomSlots = customState && customState.mode === "custom";
  const items = customState ? customState.items || [] : [];
  // click to mark loaded (ignore star clicks)
  document.querySelectorAll('.toteCard[data-idx]').forEach(el=>{
    el.addEventListener('click', (e)=>{
      if(e.target && e.target.classList && e.target.classList.contains('toteStar')) return;
      if(el.classList.contains('dragging')) return;
      const idx = parseInt(el.getAttribute('data-idx')||"0",10);
      if(!idx) return;
      toggleLoaded(routeShort, idx);
      el.classList.toggle('loaded', isLoaded(routeShort, idx));
      const r = ROUTES[activeRouteIndex];
      if(r){
        updateFooterCounts(r);
        sendMetaToParent(r);
      }
    });
  });

  // combine/uncombine
  document.querySelectorAll('.toteStar[data-action]').forEach(btn=>{
    btn.addEventListener('click', (e)=>{
      e.preventDefault(); e.stopPropagation();
      const act = btn.getAttribute('data-action');
      const second = parseInt(btn.getAttribute('data-second')||"0",10);
      if(!second) return;
      const r = ROUTES[activeRouteIndex];
      const base = baseOrder(r);

      if(act==="combine"){
        setCombined(routeShort, second, true);
      } else if(act==="uncombine"){
        setCombined(routeShort, second, false);
      }
      if(getMode(routeShort)==="custom"){
        const ord = customOrderFromSlots(routeShort, base);
        setCustomOrder(routeShort, ord);
        normalizeCustomSlots(routeShort, items);
      }
      render();
    });
  });

  // clear loaded
  const btn = document.getElementById('clearLoadedBtn');
  if(btn) btn.addEventListener('click', ()=>{ clearLoaded(routeShort); render(); });
  const rbtn = document.getElementById('resetBagsBtn');
  if(rbtn){
    rbtn.addEventListener('click', ()=>{
      const r = ROUTES[activeRouteIndex];
      const base = baseOrder(r);
      resetBagsPage(routeShort, base, items);
      if(RESET_ARMED[routeShort]){
        if(getMode(routeShort) === "custom"){
          const fallbackMode = getLastNonCustomMode(routeShort);
          const fallbackOrder = buildOrderForMode(r, fallbackMode);
          const filtered = removeCombinedSecondsFromOrder(routeShort, fallbackOrder);
          setCustomOrder(routeShort, filtered);
          setCustomSlots(routeShort, customSlotsFromOrder(filtered));
        }
        RESET_ARMED[routeShort] = false;
      }else{
        RESET_ARMED[routeShort] = true;
      }
      render();
    });
  }

  // mode buttons
  document.querySelectorAll('[data-bagmode]').forEach(b=>{
    b.addEventListener('click', ()=>{
      const nextMode = b.getAttribute('data-bagmode');
      const currentMode = getMode(routeShort);
      if(nextMode === "custom" && currentMode !== "custom"){
        setLastNonCustomMode(routeShort, currentMode);
        const r = ROUTES[activeRouteIndex];
        const base = baseOrder(r);
        const existingSlots = getCustomSlots(routeShort);
        const hasSavedCustom = existingSlots.some(slot=>slot !== null && slot !== undefined && String(slot).trim() !== "");
        if(!hasSavedCustom){
          const startingOrder = buildOrderForMode(r, currentMode);
          const filtered = removeCombinedSecondsFromOrder(routeShort, startingOrder);
          setCustomOrder(routeShort, filtered);
          setCustomSlots(routeShort, customSlotsFromOrder(filtered));
        }else{
          const filteredBase = removeCombinedSecondsFromOrder(routeShort, base);
          setCustomOrder(routeShort, customOrderFromSlots(routeShort, filteredBase));
        }
      }
      setMode(routeShort, nextMode);
      RESET_ARMED[routeShort] = false;
      render();
    });
  });

  // drag/drop for custom: swap slots
  if(!allowDrag) return;
  if(!hasCustomSlots){
    return;
  }
  let dragSlot = null;

  document.querySelectorAll('.toteCard[data-slot]').forEach(el=>{
    el.setAttribute('draggable', 'true');
    el.classList.add('draggable');

    el.addEventListener('dragstart', (e)=>{
      dragSlot = el.getAttribute('data-slot');
      el.classList.add('dragging');
      try { e.dataTransfer.setData('text/plain', dragSlot); } catch(_) {}
      e.dataTransfer.effectAllowed = 'move';
    });

    el.addEventListener('dragend', ()=>{
      dragSlot = null;
      document.querySelectorAll('[data-slot]').forEach(x=>x.classList.remove('dragging','dropTarget'));
    });
  });

  document.querySelectorAll('[data-slot]').forEach(el=>{
    el.addEventListener('dragover', (e)=>{
      e.preventDefault();
      el.classList.add('dropTarget');
      e.dataTransfer.dropEffect = 'move';
    });

    el.addEventListener('dragleave', ()=>{ el.classList.remove('dropTarget'); });

    el.addEventListener('drop', (e)=>{
      e.preventDefault();
      el.classList.remove('dropTarget');
      const targetSlot = el.getAttribute('data-slot');
      const src = dragSlot || (function(){ try { return e.dataTransfer.getData('text/plain'); } catch(_){ return null; } })();
      if(src === null || src === undefined || targetSlot === null || targetSlot === undefined) return;
      if(src === targetSlot) return;

      const from = parseInt(src, 10);
      const to = parseInt(targetSlot, 10);
      if(Number.isNaN(from) || Number.isNaN(to)) return;

      const slots = normalizeCustomSlots(routeShort, items);
      const updated = slots.slice();
      if(from < 0 || to < 0 || from >= updated.length || to >= updated.length) return;
      const fromValue = updated[from];
      if(fromValue === null || fromValue === undefined) return;
      if(from < to){
        for(let i = from; i < to; i++){
          updated[i] = updated[i + 1];
        }
        updated[to] = fromValue;
      }else if(from > to){
        for(let i = from; i > to; i--){
          updated[i] = updated[i - 1];
        }
        updated[to] = fromValue;
      }
      setCustomSlots(routeShort, updated);
      normalizeCustomSlots(routeShort, items);
      setMode(routeShort, "custom");
      clearResetArmed(routeShort);
      render();
    });
  });
}



function buildOverflowSyncOrder(r){
  const base = (r.overflow_seq || []).map((x,i)=>({
    zone: x.zone,
    count: x.count||0,
    bag_idx: x.bag_idx || 0,
    _i: i,
    _id: `${x.bag_idx||0}|${normZone(x.zone)}|${i}`,
  }));
  const buckets = new Map();
  base.forEach(item=>{
    const key = item.bag_idx || 0;
    if(!buckets.has(key)) buckets.set(key, []);
    buckets.get(key).push(item);
  });
  const ids = [];
  const seen = new Set();
  const bagOrder = buildOrder(r);
  bagOrder.forEach(idx=>{
    const items = buckets.get(idx) || [];
    items.forEach(it=>{
      if(seen.has(it._id)) return;
      seen.add(it._id);
      ids.push(it._id);
    });
    buckets.delete(idx);
  });
  base.forEach(it=>{
    if(seen.has(it._id)) return;
    seen.add(it._id);
    ids.push(it._id);
  });
  return ids;
}

function attachOverflowHandlers(routeShort, allowDrag, r){
  // mode toggle
  document.querySelectorAll('[data-ovmode]').forEach(btn=>{
    btn.addEventListener('click', ()=>{
      const m = btn.getAttribute('data-ovmode')||"normal";
      setOvMode(routeShort, m);
      if(m !== "custom") setOvOrder(routeShort, []); // keep custom order only in custom
      render();
    });
  });

  const ovSync = document.getElementById('ovSync');
  if(ovSync){
    ovSync.addEventListener('click', ()=>{
      const ids = buildOverflowSyncOrder(r);
      setOvOrder(routeShort, ids);
      setOvMode(routeShort, "custom");
      render();
    });
  }

  // checkbox toggles (click + keyboard)
  document.querySelectorAll('.ovBox[data-rowid][data-k]').forEach(box=>{
    const fire = ()=>{
      const rowId = box.getAttribute('data-rowid');
      const k = parseInt(box.getAttribute('data-k')||"0",10);
      if(!rowId || !k) return;
      toggleOvChecked(routeShort, rowId, k);
      render();
    };
    box.addEventListener('click', (e)=>{ e.preventDefault(); e.stopPropagation(); fire(); });
    box.addEventListener('keydown', (e)=>{
      if(e.key === "Enter" || e.key === " "){ e.preventDefault(); fire(); }
    });
  });

  // clear overflow checks for this route only
  const ovClear = document.getElementById('ovClear');
  if(ovClear){
    ovClear.addEventListener('click', ()=>{
      OVCHK[routeShort] = {};
      writeJSON(OVKEY, OVCHK);
      render();
    });
  }

  if(!allowDrag) return;

  // drag reorder rows (custom mode)
  let dragId = null;
  document.querySelectorAll('tr.ovDrag[data-rowid]').forEach(tr=>{
    tr.addEventListener('dragstart', (e)=>{
      dragId = tr.getAttribute('data-rowid');
      tr.classList.add('dragging');
      try{ e.dataTransfer.setData('text/plain', dragId); }catch(_){}
      e.dataTransfer.effectAllowed = 'move';
    });
    tr.addEventListener('dragend', ()=>{
      dragId = null;
      document.querySelectorAll('tr.ovDrag').forEach(x=>x.classList.remove('dragging','dropTarget'));
    });
    tr.addEventListener('dragover', (e)=>{
      e.preventDefault();
      tr.classList.add('dropTarget');
      e.dataTransfer.dropEffect = 'move';
    });
    tr.addEventListener('dragleave', ()=>tr.classList.remove('dropTarget'));
    tr.addEventListener('drop', (e)=>{
      e.preventDefault();
      tr.classList.remove('dropTarget');
      const targetId = tr.getAttribute('data-rowid');
      const srcId = dragId || (function(){ try{ return e.dataTransfer.getData('text/plain'); }catch(_){ return null; } })();
      if(!srcId || !targetId || srcId === targetId) return;

      // Build current ordered ids from DOM
      const ids = Array.from(document.querySelectorAll('tr.ovDrag[data-rowid]')).map(x=>x.getAttribute('data-rowid'));
      const from = ids.indexOf(srcId);
      const to = ids.indexOf(targetId);
      if(from === -1 || to === -1) return;
      ids.splice(from,1);
      ids.splice(to,0,srcId);
      setOvOrder(routeShort, ids);
      setOvMode(routeShort, "custom");
      render();
    });
  });
}

  function renderBags(r, q){
    const routeShort = r.route_short;
    const mode = getMode(routeShort);

    const ovMap = buildOverflowMap(r);
    const allItems = buildDisplayItems(r, "", ovMap);
    const items = q ? buildDisplayItems(r, q, ovMap) : allItems;

  function subLine(anchor, other){
    const src = anchor.sort_zone ? anchor : (other && other.sort_zone ? other : anchor);
    const sz = normZone(src.sort_zone);
    if(!sz) return "";
    return sz;
  }

  function bagBadgeText(anchor, other, idx){
    return String(anchor.idx || idx);
  }

  let layout = null;
  let slots = null;
  if(mode === "custom"){
    slots = normalizeCustomSlots(routeShort, allItems);
    const itemsById = new Map();
    items.forEach((item)=>{
      const key = slotIdForItem(item);
      if(key) itemsById.set(key, item);
    });
    layout = buildCustomSlotsLayout(routeShort, slots, itemsById, subLine, bagBadgeText, combinedPkgSum);
  }else{
    layout = buildToteLayout(items, routeShort, subLine, bagBadgeText, combinedPkgSum);
  }

  content.innerHTML = `
    <div class="toteGridFrame">
      <div class="toteWrap">
        <div class="toteBoard bagsGrid">${layout.cardsHtml}</div>
      </div>
    </div>
    <div class="bagFooter">
      <div class="bagModeDock">
        ${bagModeHtml(routeShort)}
      </div>
      <div class="footerCounts" id="footerCounts">
        <div class="countPill countPillCommercial">
          <span id="commercialCount">0</span>
          <span class="countLabel">commercial</span>
        </div>
        <div class="countPill progressPill countPillPackages">
          <span id="totalCount">0</span>
          <span class="countLabel" id="totalLabel">packages</span>
        </div>
      </div>
      <div class="clearRow">
        <button id="clearLoadedBtn" class="clearBtn">Clear</button>
        <button id="resetBagsBtn" class="clearBtn">Reset</button>
      </div>
    </div>
  `;

  const allowDrag = (mode === "custom") && !q;
  attachBagHandlers(routeShort, allowDrag, { mode, slots, items: allItems });
  updateFooterCounts(r);
  scrollTotesToRight();
}

function renderOverflow(r,q){
const routeShort = r.short || r.route_short || "";
  const mode = getOvMode(routeShort);
  const bagMeta = new Map((r.bags_detail || []).map(x=>[x.idx, x]));

  // Build ordered list (same order as Excel by default)
  const base = (r.overflow_seq || []).map((x,i)=>({
    zone: x.zone,
    count: x.count||0,
    bag_idx: x.bag_idx || 0,
    sort_zone: (bagMeta.get(x.bag_idx || 0) || {}).sort_zone || "",
    _i: i,
    _id: `${x.bag_idx||0}|${normZone(x.zone)}|${i}`
  }));

  // Apply ordering mode
  let ordered = base.slice();
  if(mode === "reversed"){
    ordered.reverse();
  } else if(mode === "custom"){
    const saved = getOvOrder(routeShort);
    if(saved && saved.length){
      const map = new Map(ordered.map(o=>[o._id,o]));
      const out = [];
      saved.forEach(id=>{ const it = map.get(id); if(it){ out.push(it); map.delete(id); } });
      // append any new rows not in saved
      map.forEach(v=>out.push(v));
      ordered = out;
    }
  }

  // Search filter keeps current order
  if(q) ordered = ordered.filter(x=>match(`${x.bag_idx} ${x.zone} ${x.count} ${x.sort_zone||""} ${normZone(x.sort_zone||"")}`, q));

  const allowDrag = (mode==="custom") && !q;

  const modeHtml = `
    <div class="ovMode" role="tablist" aria-label="Overflow order mode">
      <button class="${mode==="normal"?"on":""}" data-ovmode="normal">Normal</button>
      <button class="${mode==="reversed"?"on":""}" data-ovmode="reversed">Reversed</button>
      <button class="${mode==="custom"?"on":""}" data-ovmode="custom">Custom</button>
    </div>
  `;

  content.innerHTML = `
    <div class="ovWrap">
    <div class="ovHeader">
      <div>
        <div class="ovTitleRow">
          <button class="syncBtn" id="ovSync" type="button">Sync</button>
        </div>
      </div>
      <div class="ovHeaderRight">
        ${modeHtml}
      </div>
    </div>

    <table class="ovTable">
      <colgroup>
        <col style="width:80px">
        <col style="width:140px">
        <col>
        <col style="width:90px">
      </colgroup>
      <thead>
        <tr>
          <th style="width:70px">Bag #</th>
          <th>Overflow Zone</th>
          <th style="width:38%">Load</th>
          <th style="text-align:right;width:80px">Pkgs</th>
        </tr>
      </thead>
      <tbody>
        ${ordered.length ? ordered.map((x,idx)=>{
          const rowId = x._id;
          const total = Math.max(0, parseInt(x.count||0,10)||0);
          let done = true;
          for(let k=1;k<=total;k++){ if(!isOvChecked(routeShort,rowId,k)){ done=false; break; } }
          const trCls = `${allowDrag?'ovDrag':''} ${done && total>0 ? 'ovDone':''}`.trim();
          const checks = total ? Array.from({length: total}, (_,i)=>{
            const k=i+1;
            const on = isOvChecked(routeShort,rowId,k);
            return `<div class="ovBox ${on?'on':''}" role="checkbox" aria-checked="${on?'true':'false'}" tabindex="0" data-rowid="${rowId}" data-k="${k}"></div>`;
          }).join('') : '';
          return `
            <tr class="${trCls}" draggable="${allowDrag?'true':'false'}" data-rowid="${rowId}">
              <td style="font-weight:900">${x.bag_idx||""}</td>
              <td><span class="${normZone(x.zone).startsWith("99.") ? "ovZone ovZone99" : "ovZone"}">${normZone(x.zone)}</span></td>
              <td>${total ? `<div class="ovChecks">${checks}</div>${(done&&total>0)?`<span class="ovLoadedPill">LOADED</span>`:""}` : `<span style="color:var(--muted)">—</span>`}</td>
              <td style="text-align:right;font-weight:900">${total||""}</td>
            </tr>
          `;
        }).join("") : `<tr><td colspan="4" style="color:var(--muted)">No overflow</td></tr>`}
      </tbody>
    </table>

    <div class="rowActions">
      <button class="clearBtn" id="ovClear">Clear</button>
    </div>
  `;

  attachOverflowHandlers(routeShort, allowDrag, r);
}

function renderCombined(r,q){
  const routeShort = r.route_short;
  const mode = getMode(routeShort);
  const ovMap = buildOverflowMap(r);
  const allItems = buildDisplayItems(r, "", ovMap);
  const items = q ? buildDisplayItems(r, q, ovMap) : allItems;

  function combinedBadgeText(anchor, other){
    const src = anchor.sort_zone ? anchor : (other && other.sort_zone ? other : anchor);
    return normZone(src.sort_zone);
  }

  function combinedPkgCount(anchor, other){
    return combinedPkgSum(anchor, other);
  }

  function combinedSubLine(anchor, other){
    const first = overflowSummary(anchor.bag, ovMap);
    const second = other ? overflowSummary(other.bag, ovMap) : "";
    const parts = [first, second].filter(Boolean);
    if(!parts.length) return "";
    return parts.join("");
  }

  let layout = null;
  let slots = null;
  if(mode === "custom"){
    slots = normalizeCustomSlots(routeShort, allItems);
    const itemsById = new Map();
    items.forEach((item)=>{
      const key = slotIdForItem(item);
      if(key) itemsById.set(key, item);
    });
    layout = buildCustomSlotsLayout(routeShort, slots, itemsById, combinedSubLine, combinedBadgeText, combinedPkgCount);
  }else{
    layout = buildToteLayout(items, routeShort, combinedSubLine, combinedBadgeText, combinedPkgCount);
  }
  content.innerHTML = `
    <div class="toteGridFrame">
      <div class="toteWrap">
        <div class="toteBoard bagsGrid">${layout.cardsHtml}</div>
      </div>
    </div>
    <div class="bagFooter">
      <div class="bagModeDock">
        ${bagModeHtml(routeShort)}
      </div>
      <div class="footerCounts" id="footerCounts">
        <div class="countPill countPillCommercial">
          <span id="commercialCount">0</span>
          <span class="countLabel">commercial</span>
        </div>
        <div class="countPill progressPill countPillPackages">
          <span id="totalCount">0</span>
          <span class="countLabel" id="totalLabel">packages</span>
        </div>
      </div>
      <div class="clearRow">
        <button id="clearLoadedBtn" class="clearBtn">Clear</button>
        <button id="resetBagsBtn" class="clearBtn">Reset</button>
      </div>
    </div>
  `;

  const allowDrag = (mode === "custom") && !q;
  attachBagHandlers(routeShort, allowDrag, { mode, slots, items: allItems });
  updateFooterCounts(r);
  scrollTotesToRight();
}

let rtlScrollType = null;

function detectRtlScrollType(){
  if(rtlScrollType) return rtlScrollType;
  const probe = document.createElement("div");
  probe.dir = "rtl";
  probe.style.width = "100px";
  probe.style.height = "100px";
  probe.style.overflow = "scroll";
  probe.style.position = "absolute";
  probe.style.top = "-9999px";
  probe.style.visibility = "hidden";
  const inner = document.createElement("div");
  inner.style.width = "200px";
  inner.style.height = "1px";
  probe.appendChild(inner);
  document.body.appendChild(probe);
  probe.scrollLeft = 0;
  const start = probe.scrollLeft;
  probe.scrollLeft = 1;
  const after = probe.scrollLeft;
  document.body.removeChild(probe);
  if(start === 0 && after === 0){
    rtlScrollType = "negative";
  }else if(start === 0 && after === 1){
    rtlScrollType = "default";
  }else{
    rtlScrollType = "reverse";
  }
  return rtlScrollType;
}

function setRtlAwareScrollLeft(el, logicalLeft){
  const type = detectRtlScrollType();
  if(type === "default"){
    el.scrollLeft = logicalLeft;
    return;
  }
  if(type === "negative"){
    el.scrollLeft = -logicalLeft;
    return;
  }
  el.scrollLeft = el.scrollWidth - el.clientWidth - logicalLeft;
}

function scrollTotesToRight(){
  const wrap = document.querySelector(".toteWrap");
  if(!wrap) return;
  const maxScroll = wrap.scrollWidth - wrap.clientWidth;
  if(maxScroll > 0){
    setRtlAwareScrollLeft(wrap, maxScroll);
  }
}

function render(){
  const r = ROUTES[activeRouteIndex];
  if(!r){ content.innerHTML = "<div style='color:var(--muted)'>No routes found.</div>"; return; }
  applyWaveUI(r);
  const q = qBox.value.trim();
  content.classList.toggle('plain', activeTab==='bags' || activeTab==='combined');
  content.classList.toggle('bags-tab', activeTab==='bags');
  if(activeTab==="bags") renderBags(r,q);
  if(activeTab==="overflow") renderOverflow(r,q);
  if(activeTab==="combined") renderCombined(r,q);
  sendMetaToParent(r);
}

function buildRouteDropdown(){
  routeSel.innerHTML = "";

  const groups = new Map(); // label -> [{r,i}]
  ROUTES.forEach((r,i)=>{
    const t = (r.wave_time||"").trim();
    const wave = t ? (WAVE_LABEL_BY_TIME[t] || "Wave") : "Other";
    const label = t ? `${wave} (${t})` : "Other";
    if(!groups.has(label)) groups.set(label, []);
    groups.get(label).push({r,i});
  });

  function groupKey(label){
    const m = label.match(/^(\d+)/);
    return m ? parseInt(m[1],10) : 999;
  }
  const groupLabels = Array.from(groups.keys()).sort((a,b)=>groupKey(a)-groupKey(b) || a.localeCompare(b));

  function routeKey(rs){
    const m = rs.match(/^([A-Z]+)\.(\d+)$/);
    if(!m) return [rs, 0];
    return [m[1], parseInt(m[2],10)];
  }

  groupLabels.forEach(gl=>{
    const og = document.createElement("optgroup");
    og.label = gl;
    const arr = groups.get(gl);
    arr.sort((A,B)=>{
      const [a1,a2]=routeKey(A.r.route_short||"");
      const [b1,b2]=routeKey(B.r.route_short||"");
      if(a1!==b1) return a1.localeCompare(b1);
      return (a2-b2);
    });
    arr.forEach(({r,i})=>{
      const opt = document.createElement("option");
      opt.value = String(i);
      opt.textContent = routeTitle(r);
      og.appendChild(opt);
    });
    routeSel.appendChild(og);
  });
  adjustRouteSelectWidth();
}

function adjustRouteSelectWidth(){
  if(!routeSel) return;
  const ctx = selectMeasureCanvas.getContext("2d");
  if(!ctx) return;
  const style = getComputedStyle(routeSel);
  ctx.font = `${style.fontWeight} ${style.fontSize} ${style.fontFamily}`;
  let longest = "";
  routeSel.querySelectorAll("option, optgroup").forEach((el)=>{
    const text = (el.label || el.textContent || "").trim();
    if(text.length > longest.length) longest = text;
  });
  const textWidth = ctx.measureText(longest || "Route").width;
  const padding = parseFloat(style.paddingLeft) + parseFloat(style.paddingRight);
  const borders = parseFloat(style.borderLeftWidth) + parseFloat(style.borderRightWidth);
  const extra = 36;
  routeSel.style.width = `${Math.ceil(textWidth + padding + borders + extra)}px`;
}

function removeInternalHeaderChrome(){
  const sels = [
    "#routeTitleRow",
    "#routeTitle",
    ".routeTitleWrap",
    "#tabsRow",
    ".tabsRow",
    "#topRightTotals",
    ".topRightTotals",
    ".topCounts",
    ".topCountsExtra",
    ".sectionHeaderRow"
  ];
  sels.forEach(sel=>{
    document.querySelectorAll(sel).forEach(el=>el.remove());
  });
  ["#headerSpacer",".headerSpacer",".spacerTop"].forEach(sel=>{
    document.querySelectorAll(sel).forEach(el=>el.remove());
  });
}

function onDomReady(){
  removeInternalHeaderChrome();
  const main = document.querySelector("#main")
    || document.querySelector(".main")
    || document.querySelector(".organizerBody")
    || document.body;
  if(main){
    main.style.paddingTop = "0px";
    main.style.marginTop = "0px";
  }
}

if(document.readyState === "loading"){
  document.addEventListener("DOMContentLoaded", onDomReady);
}else{
  onDomReady();
}

function init(){
  buildRouteDropdown();
  if(routeSel){
    routeSel.value = String(activeRouteIndex);
  }
  routeSel.addEventListener("change", ()=>{ activeRouteIndex=parseInt(routeSel.value,10)||0; render(); });

  fetch("toc-data", { cache:"no-store" })
    .then(r=>r.json())
    .then(data=>{
      if(data && data.status==="ok"){
        WAVE_COLORS = data.wave_colors || {};
        rebuildDropdownWithWaveDots();
        applyWaveUI(ROUTES[activeRouteIndex]);
      }
    })
    .catch(()=>{});

  qBox.addEventListener("input", ()=>render());
  if(organizerRoot && "ResizeObserver" in window){
    const ro = new ResizeObserver(()=>{
      scheduleRender();
    });
    ro.observe(organizerRoot);
  }
  if(document.fonts && document.fonts.ready){
    document.fonts.ready.then(()=>{
      scheduleRender();
    });
  }
  window.addEventListener('orientationchange', ()=>{
    scheduleRender();
  });
  window.addEventListener('resize', ()=>{
    scheduleRender();
    adjustRouteSelectWidth();
  });
  render();
}
init();

window.addEventListener("message", (ev)=>{
  const d = ev.data || {};
  if(d.type !== "setTab") return;
  if(d.tab === "bags_overflow") setActiveTab("combined");
  if(d.tab === "bags") setActiveTab("bags");
  if(d.tab === "overflow") setActiveTab("overflow");
});
</script>
<script>
(function(){
  function fitToteGridToFrame(){
    var frame = document.querySelector('.toteGridFrame');
    var wrap = frame && frame.querySelector('.toteWrap');
    var grid = wrap && wrap.querySelector('.bagsGrid');
    if(!frame || !wrap || !grid) return;

    var cards = Array.from(grid.children).filter(function(el){
      return el.classList && el.classList.contains('toteCard');
    });
    var total = grid.children.length || cards.length || 0;
    var wrapRect = wrap.getBoundingClientRect();
    var availW = Math.max(0, wrapRect.width);
    var availH = Math.max(0, wrapRect.height);
    if(!availW || !availH) return;

    var gridStyle = getComputedStyle(grid);
    var padX = (parseFloat(gridStyle.paddingLeft) || 0) + (parseFloat(gridStyle.paddingRight) || 0);
    var padY = (parseFloat(gridStyle.paddingTop) || 0) + (parseFloat(gridStyle.paddingBottom) || 0);
    var innerW = Math.max(0, availW - padX);
    var innerH = Math.max(0, availH - padY);
    if(!innerW || !innerH) return;

    var baseW = parseFloat(gridStyle.getPropertyValue('--tote-base-w')) || 210;
    var baseH = parseFloat(gridStyle.getPropertyValue('--tote-base-h')) || 190;
    var minScale = parseFloat(gridStyle.getPropertyValue('--tote-min-scale')) || 0.55;
    var maxScale = parseFloat(gridStyle.getPropertyValue('--tote-max-scale')) || 1.15;
    var gapX = parseFloat(gridStyle.columnGap || gridStyle.gap) || 0;
    var gapY = parseFloat(gridStyle.rowGap || gridStyle.gap) || 0;
    var card = cards[0];
    var cardStyle = card ? getComputedStyle(card) : null;
    var cardPadW = cardStyle
      ? (parseFloat(cardStyle.paddingLeft) || 0) + (parseFloat(cardStyle.paddingRight) || 0)
      : 0;
    var cardPadH = cardStyle
      ? (parseFloat(cardStyle.paddingTop) || 0) + (parseFloat(cardStyle.paddingBottom) || 0)
      : 0;
    var rows = 3;
    var cols = Math.max(1, Math.ceil(total / rows));
    var totalGapX = gapX * Math.max(0, cols - 1);
    var totalGapY = gapY * Math.max(0, rows - 1);
    var cellW = (innerW - totalGapX) / cols;
    var cellH = (innerH - totalGapY) / rows;
    if(cellW <= 0 || cellH <= 0) return;

    var contentW = cellW - cardPadW;
    var contentH = cellH - cardPadH;
    if(contentW <= 0 || contentH <= 0) return;
    var isNarrow = window.matchMedia && window.matchMedia("(max-width: 900px)").matches;
    var rawScale = isNarrow ? (contentH / baseH) : Math.min(contentW / baseW, contentH / baseH);
    var scale = Math.min(maxScale, Math.max(minScale, rawScale));
    if(isNarrow){
      var minCellW = (baseW * scale) + cardPadW;
      grid.style.setProperty('--tote-min-cell-w', Math.ceil(minCellW) + 'px');
    }

    grid.style.setProperty('--tote-rows', rows);
    grid.style.setProperty('--tote-cols', cols);
    grid.style.setProperty('--tote-scale', scale.toFixed(3));
  }

  var _fitTimer = null;
  function refitSoon(){
    clearTimeout(_fitTimer);
    _fitTimer = setTimeout(fitToteGridToFrame, 60);
  }

  // Hook into existing render if present
  var _render = window.render;
  if(typeof _render === 'function'){
    window.render = function(){
      var out = _render.apply(this, arguments);
      fitToteGridToFrame();
      return out;
    };
  }

  window.addEventListener('load', fitToteGridToFrame);
  window.addEventListener('resize', refitSoon);
  if (window.visualViewport){
    visualViewport.addEventListener('resize', refitSoon);
    visualViewport.addEventListener('scroll', refitSoon);
  }

  if(window.ResizeObserver){
    var frame = document.querySelector('.toteGridFrame');
    if(frame){
      var ro = new ResizeObserver(refitSoon);
      ro.observe(frame);
    }
  }

  // Refit on DOM mutations (route switch, tab change, etc.)
  if(window.MutationObserver){
    var target = document.querySelector('.toteGridFrame') || document.body;
    var mo = new MutationObserver(refitSoon);
    mo.observe(target, { childList:true, subtree:true, attributes:true });
  }
})();
</script>
</body>
</html>
"""


def build_html(header_title: str, routes: List[dict], wave_map: dict) -> str:
    # Keep JSON dumps settings identical to previous: no indent, ensure_ascii False.
    routes_json = json.dumps(routes, ensure_ascii=False)
    wave_json = json.dumps(wave_map, ensure_ascii=False)
    route_code = header_title
    route_date = ""
    route_sep = ""
    if " • " in header_title:
        route_code, route_date = header_title.split(" • ", 1)
        route_sep = " • "
    return (HTML_TEMPLATE
            .replace("__HEADER_TITLE__", header_title)
            .replace("__HEADER_ROUTE_CODE__", route_code)
            .replace("__HEADER_ROUTE_DATE__", route_date)
            .replace("__HEADER_ROUTE_SEP__", route_sep)
            .replace("__ROUTES_JSON__", routes_json)
            .replace("__WAVE_JSON__", wave_json))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--pdf", required=True, help="Route Sheets PDF")
    ap.add_argument("--xlsx", required=True, help="Bags_with_Overflow Excel")
    ap.add_argument("--out", required=True, help="Output HTML path")
    ap.add_argument("--no-cache", action="store_true", help="Disable PDF parse cache")
    args = ap.parse_args()

    header_title, _, pdf_meta, route_time, pkg_summary = parse_pdf_meta(
        args.pdf,
        use_cache=not args.no_cache,
    )
    if args.no_cache:
        routes = parse_excel_routes(args.xlsx, pdf_meta, route_time, pkg_summary)
        wave_map = build_wave_labels(routes)
    else:
        cached = _load_routes_cache(args.pdf, args.xlsx)
        if cached:
            routes = cached["routes"]
            wave_map = cached["wave_map"]
        else:
            routes = parse_excel_routes(args.xlsx, pdf_meta, route_time, pkg_summary)
            wave_map = build_wave_labels(routes)
            _save_routes_cache(args.pdf, args.xlsx, {"routes": routes, "wave_map": wave_map})
    html = build_html(header_title, routes, wave_map)
    Path(args.out).write_text(html, encoding="utf-8")
    print(args.out)


if __name__ == "__main__":
    main()
